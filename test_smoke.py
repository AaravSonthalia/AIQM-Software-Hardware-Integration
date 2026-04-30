"""
Smoke test for Growth Monitor session lifecycle.

Exercises GrowthLogger directly (no GUI needed) to verify all outputs:
  - sensor_log.csv
  - commit_log.csv
  - frames/ directory with PNG files
  - session_metadata.json
  - growth_log.xlsx (or CSV fallback)

Run: python test_smoke.py
"""

import json
import csv
import shutil
from datetime import datetime
from pathlib import Path

import numpy as np

from gui.growth_logger import GrowthLogger


TEST_DIR = "logs/_smoke_test"
PASS = "\033[92mPASS\033[0m"
FAIL = "\033[91mFAIL\033[0m"


def make_fake_frame(width: int = 656, height: int = 492) -> np.ndarray:
    """Create a synthetic RGB frame resembling a RHEED pattern."""
    frame = np.zeros((height, width, 3), dtype=np.uint8)
    # Green-channel Gaussian blob (mimics RHEED spot)
    y, x = np.ogrid[:height, :width]
    cx, cy = width // 2, height // 2
    r2 = (x - cx) ** 2 + (y - cy) ** 2
    frame[:, :, 1] = (200 * np.exp(-r2 / (2 * 80**2))).astype(np.uint8)
    return frame


def check(label: str, condition: bool, detail: str = ""):
    status = PASS if condition else FAIL
    msg = f"  [{status}] {label}"
    if detail and not condition:
        msg += f"  — {detail}"
    print(msg)
    return condition


def run_smoke_test():
    print("=" * 60)
    print("Growth Monitor — Smoke Test")
    print("=" * 60)

    # Clean up any previous test run
    test_base = Path(TEST_DIR)
    if test_base.exists():
        shutil.rmtree(test_base)

    logger = GrowthLogger(base_dir=TEST_DIR)
    all_passed = True

    # --- 1. Start session ---
    print("\n1. Starting session...")
    logger.start_session("STO_TEST_001")
    session_dir = logger.session_dir
    check("Session dir created", session_dir is not None and session_dir.exists())
    check("Session dir name format",
          "STO_TEST_001" in str(session_dir),
          f"Got: {session_dir}")
    check("frames/ subdir exists", (session_dir / "frames").exists())
    check("sensor_log.csv created", (session_dir / "sensor_log.csv").exists())
    check("commit_log.csv created", (session_dir / "commit_log.csv").exists())
    check("Logger is active", logger.active)

    # --- 2. Log sensor readings ---
    print("\n2. Logging sensor data...")
    for i in range(5):
        logger.log_sensors(pyro_temp=450.0 + i * 10, elapsed_s=float(i))

    with open(session_dir / "sensor_log.csv") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    check("sensor_log.csv has 5 rows", len(rows) == 5, f"Got {len(rows)}")
    check("sensor_log has timestamp field", "timestamp" in rows[0])
    check("sensor_log has elapsed_s field", "elapsed_s" in rows[0])
    check("sensor_log has pyrometer_temp_C field", "pyrometer_temp_C" in rows[0])
    check("First temp is 450.0", rows[0]["pyrometer_temp_C"] == "450.0",
          f"Got: {rows[0]['pyrometer_temp_C']}")

    # --- 3. Commit log entries with frames ---
    print("\n3. Committing log entries with RHEED frames...")
    recon_samples = [
        {"recon_1x1": "80", "recon_Twinned (2x1)": "10", "recon_c(6x2)": "5",
         "recon_rt13xrt13": "3", "recon_HTR": "2"},
        {"recon_1x1": "0", "recon_Twinned (2x1)": "0", "recon_c(6x2)": "60",
         "recon_rt13xrt13": "30", "recon_HTR": "10"},
        {"recon_1x1": "0", "recon_Twinned (2x1)": "0", "recon_c(6x2)": "0",
         "recon_rt13xrt13": "5", "recon_HTR": "95"},
    ]
    for i in range(3):
        frame = make_fake_frame()
        ts = f"14{i:02d}00"
        path = logger.save_frame(frame, ts)
        entry = {
            "timestamp": datetime.now().isoformat(),
            "time_display": f"14:{i:02d}",
            "sample_id": "STO_TEST_001",
            "grower": "TestGrower",
            "elapsed_s": f"{i * 60:.2f}",
            "pyrometer_temp_C": f"{500.0 + i * 50:.1f}",
            "voltage_V": "",
            "current_A": "",
            **recon_samples[i],
            "note": f"Test note #{i + 1}",
            "frame_path": path,
        }
        logger.log_commit(entry)

    # Verify frames
    frames_dir = session_dir / "frames"
    bmps = sorted(frames_dir.glob("*.bmp"))
    check("3 BMP frames saved", len(bmps) == 3, f"Found {len(bmps)}: {[p.name for p in bmps]}")
    for bmp in bmps:
        check(f"  Frame {bmp.name} > 0 bytes", bmp.stat().st_size > 0,
              f"Size: {bmp.stat().st_size}")

    # Verify frame is valid image
    from PIL import Image
    if bmps:
        img = Image.open(bmps[0])
        check(f"  First frame dimensions correct",
              img.size == (656, 492),
              f"Got: {img.size}")
        check(f"  First frame is BMP format",
              img.format == "BMP",
              f"Got: {img.format}")

    # Verify commit log
    with open(session_dir / "commit_log.csv") as f:
        reader = csv.DictReader(f)
        commits = list(reader)
    check("commit_log.csv has 3 rows", len(commits) == 3, f"Got {len(commits)}")
    check("commit_log has frame_path", "frame_path" in commits[0])
    check("commit_log frame_path points to file",
          Path(commits[0]["frame_path"]).exists(),
          f"Path: {commits[0]['frame_path']}")
    check("commit_log has note", commits[0]["note"] == "Test note #1",
          f"Got: {commits[0]['note']}")
    check("commit_log has recon_1x1 field", "recon_1x1" in commits[0])
    check("commit_log has recon_HTR field", "recon_HTR" in commits[0])
    check("First entry recon_1x1 == 80", commits[0]["recon_1x1"] == "80",
          f"Got: {commits[0].get('recon_1x1')}")
    check("Third entry recon_HTR == 95", commits[2]["recon_HTR"] == "95",
          f"Got: {commits[2].get('recon_HTR')}")

    # --- 4. Save metadata & export ---
    print("\n4. Stopping session — saving metadata and exporting...")
    metadata = {
        "date": datetime.now().strftime("%Y-%m-%d"),
        "grower": "TestGrower",
        "sample_id": "STO_TEST_001",
    }
    logger.save_session_metadata(metadata)

    # Verify metadata
    meta_path = session_dir / "session_metadata.json"
    check("session_metadata.json created", meta_path.exists())
    with open(meta_path) as f:
        meta = json.load(f)
    check("metadata has session_end", "session_end" in meta)
    check("metadata total_entries == 3", meta.get("total_entries") == 3,
          f"Got: {meta.get('total_entries')}")
    check("metadata has grower", meta.get("grower") == "TestGrower")
    check("metadata has sample_id", meta.get("sample_id") == "STO_TEST_001")

    # Export growth log
    export_path = logger.export_growth_log(metadata)
    check("Growth log exported", export_path != "")

    if export_path.endswith(".xlsx"):
        check("Export is xlsx format", True)
        from openpyxl import load_workbook
        wb = load_workbook(export_path)
        ws = wb.active
        check("xlsx has 'Growth Log' sheet", ws.title == "Growth Log")
        check("xlsx has grower", ws['D3'].value == "TestGrower")
        check("xlsx has sample_id", ws['I3'].value == "STO_TEST_001")
        # Check operations rows start at row 13
        check("xlsx row 13 has first entry time",
              ws.cell(row=13, column=2).value == "14:00",
              f"Got: {ws.cell(row=13, column=2).value}")
        check("xlsx row 13 has first entry note",
              ws.cell(row=13, column=4).value == "Test note #1",
              f"Got: {ws.cell(row=13, column=4).value}")
        check("xlsx row 13 has temperature",
              ws.cell(row=13, column=3).value == 500.0,
              f"Got: {ws.cell(row=13, column=3).value}")
    else:
        check("Export is csv fallback", export_path.endswith(".csv"))

    # End session
    logger.end_session()
    check("Logger no longer active after end_session", not logger.active)
    check("Session dir preserved after end", logger.session_dir is not None)

    # --- 5. Verify complete session directory ---
    print("\n5. Final session directory contents:")
    for item in sorted(session_dir.rglob("*")):
        if item.is_file():
            size = item.stat().st_size
            rel = item.relative_to(session_dir)
            print(f"  {rel}  ({size:,} bytes)")

    # --- Summary ---
    print("\n" + "=" * 60)
    print("Smoke test complete.")
    print(f"Session directory: {session_dir}")
    print("=" * 60)

    # Cleanup prompt
    print(f"\nTo clean up: rm -rf {TEST_DIR}")


if __name__ == "__main__":
    run_smoke_test()
