"""
Session data logger for MBE growth monitoring.

Creates a session directory with periodic sensor CSV, commit log CSV,
saved RHEED frames, session metadata JSON, and growth log export.
"""

import csv
import hashlib
import json
import logging
import math
import os
import tempfile
import uuid
from collections.abc import Mapping
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable, Optional

import numpy as np

from .equalizer_alignment import (
    BasisBundle,
    CalibrationRecord,
    RheedFrameSnapshot,
    normalize_utc_timestamp,
)
from .equalizer_label_contract import (
    ACTIVE_LABELS as EQUALIZER_ACTIVE_LABELS,
    CSV_SUFFIX as EQUALIZER_CSV_SUFFIX,
    frame_rgb_sha256,
    validate_equalizer_payload,
)


log = logging.getLogger(__name__)


# Auto-capture event states (used by event_state column in
# auto_capture_events.csv). Set to PENDING at fire time; transitions to one
# of the other three when the grower interacts with the AutoCaptureBanner or
# the keep-default countdown fires. AUTO_SKIPPED is set at log time when
# the quality gate rejected all 20 buffer frames — there's nothing to
# review, so the event has no banner and no resolution path. Marking it
# at log time gives the row a terminal state instead of leaving it as
# pending forever.
EVENT_STATE_PENDING = "pending"
EVENT_STATE_KEPT_EXPLICIT = "kept_explicit"
EVENT_STATE_KEPT_DEFAULT = "kept_default"
EVENT_STATE_DISCARDED = "discarded"
EVENT_STATE_AUTO_SKIPPED = "auto_skipped"


@dataclass(frozen=True)
class AutoCaptureCommitResult:
    """Durable auto-event outcome, separate from append-stream readiness."""

    buffer_count: int = 0
    buffer_dir: str = ""
    committed: bool = False
    writer_ready: bool = False

# Structured RHEED view/QC event types. These are deliberately separate
# from ``manual_events.csv``: reconstruction-change marks and acquisition
# validity transitions have different semantics and downstream consumers.
RHEED_VIEW_EVENT_SESSION_START = "session_start"
RHEED_VIEW_EVENT_ALIGNMENT_CONFIRMED = "alignment_confirmed"
RHEED_VIEW_EVENT_REALIGN_START = "realign_start"
RHEED_VIEW_EVENT_REALIGN_END = "realign_end"
RHEED_VIEW_EVENT_HISTORY_RESET = "history_reset"
RHEED_VIEW_EVENT_HISTORY_READY = "history_ready"
RHEED_VIEW_EVENT_QC_REJECT = "qc_reject"
RHEED_VIEW_EVENT_QC_PASS = "qc_pass"
RHEED_VIEW_EVENT_TYPES = frozenset({
    RHEED_VIEW_EVENT_SESSION_START,
    RHEED_VIEW_EVENT_ALIGNMENT_CONFIRMED,
    RHEED_VIEW_EVENT_REALIGN_START,
    RHEED_VIEW_EVENT_REALIGN_END,
    RHEED_VIEW_EVENT_HISTORY_RESET,
    RHEED_VIEW_EVENT_HISTORY_READY,
    RHEED_VIEW_EVENT_QC_REJECT,
    RHEED_VIEW_EVENT_QC_PASS,
})


class GrowthLogger:
    """Logs sensor data and timestamped entries during a growth session."""

    LIVE_LABEL_TRANSACTION_FILE = ".live_label_transaction.pending.json"
    LIVE_LABEL_TRANSACTION_SCHEMA_VERSION = 1
    CALIBRATION_TRANSACTION_FILE = ".equalizer_calibration.pending.json"
    CALIBRATION_TRANSACTION_SCHEMA_VERSION = 1
    AUTO_CAPTURE_TRANSACTION_FILE = ".auto_capture_transaction.pending.json"
    AUTO_CAPTURE_TRANSACTION_SCHEMA_VERSION = 2
    HUMAN_LABELING_STATE_FILE = "human_labeling_state.json"
    HUMAN_LABELING_STATE_SCHEMA_VERSION = 1

    # Source of truth for independent human primary judgments.  Unlike
    # events_labels.csv this file is never rewritten or keyed by event_idx:
    # every click is a distinct annotation that can participate in a vote.
    HUMAN_PRIMARY_LABEL_FIELDS = [
        "schema_version", "annotation_id", "label_idx",
        "label_timestamp_iso", "session_id", "acquisition_run",
        "frame_path", "frame_sha256", "frame_sha256_algorithm",
        "capture_sequence", "view_segment_id", "capture_backend",
        "captured_at_utc", "gun_aligned", "realignment_active",
        "human_primary_reconstruction", "human_label_source",
        "human_labeler", "human_confidence",
        "human_blind_to_model", "human_blind_to_equalizer",
        "qc_label",
    ]

    SENSOR_FIELDS = [
        "timestamp", "elapsed_s",
        "pyrometer_temp_C", "pyrometer_temp_std_C", "pyrometer_temp_n",
        "mistral_v_set_V", "mistral_v_actual_V",
        "mistral_i_set_A", "mistral_i_actual_A",
        "chamber_pressure_mbar",
        # Elog-direct columns (populated when EvapControl mode = "elog";
        # blank in screengrab mode where the .elo binary isn't read).
        # See drivers.evap_control.ElogReader.DEFAULT_VAR_MAP for the
        # elog-variable-name → column-name mapping.
        "substrate_temp_pv_C", "substrate_temp_setpoint_C",
        "cell_HTEC2_pv_C",
        "cell_Y_pv_C", "cell_Sr_pv_C", "cell_Eu_pv_C", "cell_Er_pv_C",
        "plasma_dc_bias_V", "plasma_forward_W", "plasma_reflected_W",
        # ADS-mode cell temperatures (Ch-MBE, mode="ads"). Empty in O-MBE
        # sessions (MistralState.ads_cells is None). Cell1–7 map to the
        # Beckhoff PLC's PIDProgram.Cell{i}_pidTDK.ActualTemperature.
        "cell1_T_C", "cell2_T_C", "cell3_T_C", "cell4_T_C",
        "cell5_T_C", "cell6_T_C", "cell7_T_C",
    ]
    COMMIT_FIELDS = [
        "timestamp", "time_display", "elapsed_s", "sample_id", "grower",
        "pyrometer_temp_C", "voltage_V", "current_A",
        # Source path that populated voltage_V/current_A. Named for the
        # method, not the vendor, so it stays correct through hardware
        # swaps. Three values:
        #   mistral — via MISTRAL's screengrab OCR (current O-MBE path;
        #             TDK-Lambda hardware → MISTRAL software → GUI OCR)
        #   direct  — via a direct-read PSU worker (currently unwired;
        #             future TDK/OWON/etc. when we get a serial or
        #             network path to the PSU without going through
        #             MISTRAL)
        #   none    — no PSU state cached at LOG time (voltage_V and
        #             current_A will be blank)
        # Downstream analysis (Yuxin's #1) reads this to know whether
        # a voltage_V value came from screengrab OCR (~1 Hz, subject to
        # OCR noise) or a direct instrument read (higher precision).
        "psu_source",
        # Slider values at LOG time. When ``grower_corrected == False`` these
        # equal the classifier's smoothed_percent (sliders are read-only and
        # mirror the classifier). When ``grower_corrected == True`` these
        # represent the grower's belief, sum to 100 (Pattern A proportional
        # adjustment enforces it), and can differ from ``classifier_recon_*``.
        "recon_1x1", "recon_Twinned (2x1)", "recon_c(6x2)",
        "recon_rt13xrt13", "recon_HTR",
        "recon_label_source", "recon_labeler", "recon_confidence",
        # Classifier's smoothed_percent at LOG time — always populated when a
        # classifier state has been received this session, regardless of
        # correction toggle. Empty when the classifier is disabled or the
        # session logged before any frame was classified. Pairs with the
        # matching ``recon_*`` columns to form the active-comparisons training
        # signal for Yuxin's #1 deliverable (see yuxin_deliverables_jul06.md).
        "classifier_recon_1x1", "classifier_recon_Twinned (2x1)",
        "classifier_recon_c(6x2)", "classifier_recon_rt13xrt13",
        "classifier_recon_HTR",
        # "True" when the grower had ``✎ Correct`` active at LOG time (may
        # equal classifier values if grower agreed without dragging), "False"
        # when correction was off, "" when the classifier itself was disabled
        # for the session.
        "grower_corrected",
        # Classifier lifecycle state at LOG time. Four values:
        #   OK       — classifier inference ran; use the input-mode and
        #              actionable columns for deployment semantics
        #   LOADING  — worker up but no successful classify() yet
        #   ERROR    — worker emitted an error (missing model, bad state)
        #   DISABLED — classifier never armed for this session (config off)
        # Disambiguates the classifier_recon_* columns downstream. A "0"
        # in classifier_recon_1x1 could mean "classifier confident it's
        # not 1x1" (OK), "classifier initialising" (LOADING), or "no
        # data" (blank when ERROR/DISABLED, but the status column makes
        # the reason explicit). Yuxin's #1 pipeline reads this to
        # partition rows appropriately.
        "classifier_status",
        # Runtime semantics needed to distinguish a successful single-frame
        # research output from a future temporal/actionable adviser.
        "classifier_input_mode", "classifier_prediction_actionable",
        "classifier_view_segment_id",
        "classifier_visual_history_generation",
        "classifier_history_ready",
        "note", "frame_path",
        # Boolean-ish (str "True"/"False"/""). "True" when the saved
        # frame passed the frame_quality gate at LOG time; "False" when
        # the gate flagged it (too dark, uniform, saturated) but LOG
        # ENTRY saved it anyway because the grower explicitly requested
        # the entry. Empty when no frame was available at all (camera
        # not connected / no state emitted yet). Downstream can filter
        # or weight rows by this — training data curators may want
        # only frame_quality_pass=True rows, while UX audits may want
        # every entry.
        "frame_quality_pass",
        "capture_backend", "captured_at_utc", "capture_sequence",
        "frame_age_ms", "source_hwnd", "capture_geometry_id",
    ]
    AUTO_CAPTURE_FIELDS = [
        "timestamp", "elapsed_s", "event_idx",
        "change_score", "pyrometer_temp_C",
        "buffer_count", "buffer_dir",
        "event_state", "state_changed_at",
        "capture_backend", "captured_at_utc", "capture_sequence",
        "frame_age_ms", "source_hwnd", "capture_geometry_id",
    ]
    HEARTBEAT_FIELDS = [
        "timestamp", "elapsed_s", "heartbeat_idx",
        "pyrometer_temp_C", "frame_path",
        "capture_backend", "captured_at_utc", "capture_sequence",
        "frame_age_ms", "source_hwnd", "capture_geometry_id",
    ]
    SET_CHANGE_FIELDS = [
        "timestamp", "elapsed_s", "event_idx",
        "channel", "old_value", "new_value", "delta",
        "pyrometer_temp_C",
    ]
    # Grower-marked events (Jul 10 2026 group-meeting design shift). One
    # row per MARK EVENT click. Deliberately kept lean so the click stays
    # sub-second — no classifier snapshot, no slider dump. The scrubber
    # (workstream #3) reads this file to render manual markers on the
    # timeline; the wider (classifier + slider) capture stays on LOG
    # ENTRY where the grower opts in to the fuller decision. See
    # meeting_jul10_2026.md for the three-concern architecture
    # (capture / mark / label) that motivated the split from
    # AUTO_CAPTURE_FIELDS.
    MANUAL_EVENT_FIELDS = [
        "timestamp", "elapsed_s", "event_idx",
        "pyrometer_temp_C",
        "voltage_V", "current_A", "psu_source",
        "frame_path",
        # Optional — populated from the log-note input if the grower had
        # text queued at click time. Blank otherwise (single-tap case is
        # the primary path).
        "note",
        "capture_backend", "captured_at_utc", "capture_sequence",
        "frame_age_ms", "source_hwnd", "capture_geometry_id",
    ]
    # Acquisition-geometry and global image-QC transitions. State columns
    # describe the post-event state; ``frame_role`` identifies whether the
    # optional snapshot is the pre- or post-realignment view. Keeping this
    # in its own CSV preserves the existing manual-event schema and prevents
    # global QC_REJECT from being confused with reconstruction-specific
    # not_apply labels.
    RHEED_VIEW_EVENT_FIELDS = [
        "timestamp", "elapsed_s", "event_idx", "event_type",
        "realignment_id",
        "previous_view_segment_id", "view_segment_id",
        "visual_history_generation",
        "gun_aligned",
        "history_frame_count", "history_required", "history_ready",
        "qc_reject", "qc_reason", "labeler", "confidence",
        "prediction_actionable",
        "frame_role", "frame_path", "note",
        "capture_backend", "captured_at_utc", "capture_sequence",
        "frame_age_ms", "source_hwnd", "capture_geometry_id",
    ]
    # Live-Equalizer labels (Jul 10 2026 group-meeting workstream #4).
    # One row per Save click in the Live Equalizer window. The 5 recon_*
    # columns are the sliders' current mixture (Pattern A normalization
    # NOT applied here; grower chooses whether to hit Normalize before
    # Save). Coexists with commit_log.csv:recon_* (which are the Monitor-
    # tab sliders on a LOG ENTRY) — same class labels, different capture
    # surface + rate. Downstream analyses can join on timestamp.
    LIVE_LABEL_FIELDS = [
        "timestamp", "elapsed_s", "label_idx",
        "equalizer_schema_version", "equalizer_source",
        "equalizer_labeler", "equalizer_confidence", "equalizer_argmax",
        "equalizer_fit_mode", "equalizer_normalization_applied",
        "equalizer_final_is_normalized", "equalizer_final_sum",
        "equalizer_fit_residual", "equalizer_valid_coverage",
        "equalizer_calibration_valid",
        *[f"equalizer_raw_{suffix}" for suffix in EQUALIZER_CSV_SUFFIX.values()],
        "equalizer_raw_HTR",
        *[f"equalizer_final_{suffix}" for suffix in EQUALIZER_CSV_SUFFIX.values()],
        "equalizer_final_HTR",
        *[
            f"equalizer_normalized_{suffix}"
            for suffix in EQUALIZER_CSV_SUFFIX.values()
        ],
        "equalizer_normalized_HTR",
        "recon_1x1", "recon_tw", "recon_c6x2",
        "recon_rt13", "recon_HTR",
        "pyrometer_temp_C",
        "voltage_V", "current_A", "psu_source",
        "frame_path",
        "capture_backend", "captured_at_utc", "capture_sequence",
        "captured_monotonic_ns", "frame_age_ms", "source_hwnd",
        "capture_geometry_id", "session_id", "gun_aligned",
        "realignment_active",
        "calibration_id", "basis_bundle_id", "equalizer_active_classes",
        "view_segment_id", "visual_history_generation",
        "equalizer_frame_sha256", "equalizer_frame_sha256_algorithm",
    ]
    # Event labels written by the Events tab labeling form. The from/to
    # columns are reserved for the deferred reconstruction-transition
    # dropdowns — they exist now so future UI additions don't require a
    # CSV schema migration. Each row is per-event_idx (upsert-by-key).
    EVENT_LABEL_FIELDS = [
        "event_idx",
        "primary_reconstruction",
        "human_primary_reconstruction", "human_label_source",
        "human_labeler", "human_confidence",
        "human_blind_to_model", "human_blind_to_equalizer",
        "human_primary_annotation_id", "human_primary_label_idx",
        "human_primary_frame_path", "human_primary_frame_sha256",
        "human_primary_frame_sha256_algorithm",
        "human_primary_capture_backend", "human_primary_captured_at_utc",
        "human_primary_capture_sequence", "human_primary_view_segment_id",
        "human_primary_visual_history_generation", "human_primary_session_id",
        "change_from",
        "change_to",
        "notes",
        "label_timestamp_iso",
        # Equalizer is an independent visual-basis measurement.  Its argmax
        # and weights must never populate either human-primary column.
        "equalizer_schema_version", "equalizer_source",
        "equalizer_labeler", "equalizer_confidence", "equalizer_argmax",
        "equalizer_fit_mode", "equalizer_normalization_applied",
        "equalizer_final_is_normalized", "equalizer_final_sum",
        "equalizer_fit_residual", "equalizer_valid_coverage",
        "equalizer_calibration_valid",
        *[f"equalizer_raw_{suffix}" for suffix in EQUALIZER_CSV_SUFFIX.values()],
        "equalizer_raw_HTR",
        *[f"equalizer_final_{suffix}" for suffix in EQUALIZER_CSV_SUFFIX.values()],
        "equalizer_final_HTR",
        *[
            f"equalizer_normalized_{suffix}"
            for suffix in EQUALIZER_CSV_SUFFIX.values()
        ],
        "equalizer_normalized_HTR",
        "recon_1x1",
        "recon_tw",
        "recon_c6x2",
        "recon_rt13",
        "recon_HTR",
        "calibration_id", "basis_bundle_id", "equalizer_active_classes",
        "view_segment_id", "visual_history_generation",
        "frame_path",
        "capture_backend", "captured_at_utc", "capture_sequence",
        "frame_age_ms", "source_hwnd", "capture_geometry_id",
        "captured_monotonic_ns",
        "gun_aligned", "realignment_active",
        "session_id",
        "equalizer_frame_sha256", "equalizer_frame_sha256_algorithm",
    ]

    AUTO_CAPTURE_MANIFEST_FIELDS = [
        "frame_path",
        "capture_backend", "captured_at_utc", "capture_sequence",
        "captured_monotonic_ns", "frame_age_ms", "source_hwnd",
        "capture_geometry_id",
        "camera_width", "camera_height",
        "session_id", "view_segment_id", "visual_history_generation",
        "gun_aligned", "realignment_active", "calibration_id",
        "basis_bundle_id",
    ]

    @staticmethod
    def _capture_columns(capture_metadata: Optional[dict] = None) -> dict:
        """Normalize optional camera provenance for CSV writers."""
        metadata = capture_metadata or {}
        frame_age = metadata.get("frame_age_ms", "")
        if isinstance(frame_age, (float, int)):
            frame_age = f"{float(frame_age):.3f}"
        return {
            "capture_backend": metadata.get("capture_backend", ""),
            "captured_at_utc": metadata.get("captured_at_utc", ""),
            "capture_sequence": metadata.get("capture_sequence", ""),
            "frame_age_ms": frame_age,
            "source_hwnd": metadata.get("source_hwnd", ""),
            "capture_geometry_id": metadata.get("capture_geometry_id", ""),
        }

    @staticmethod
    def _active_classes_text(active_classes: Iterable[str]) -> str:
        """Serialize active classes deterministically for CSV consumers."""
        if isinstance(active_classes, str):
            try:
                decoded = json.loads(active_classes)
            except json.JSONDecodeError:
                decoded = [part.strip() for part in active_classes.split(",")]
            if not isinstance(decoded, list):
                raise ValueError("Equalizer active classes must be a sequence")
            values = tuple(str(label) for label in decoded if str(label))
        else:
            values = tuple(str(label) for label in active_classes)
        if not values or len(set(values)) != len(values):
            raise ValueError("Equalizer active classes must be unique and non-empty")
        return json.dumps(values, separators=(",", ":"))

    @staticmethod
    def _equalizer_record_columns(
        payload: Optional[Mapping[str, object]],
        *,
        labeler: str,
        rgb: Optional[np.ndarray],
        calibration_valid: bool,
    ) -> dict[str, object]:
        """Flatten a validated measurement payload for either Equalizer CSV."""
        columns: dict[str, object] = {
            "equalizer_schema_version": "",
            "equalizer_source": "legacy_unqualified",
            "equalizer_labeler": str(labeler or "").strip(),
            "equalizer_confidence": "",
            "equalizer_argmax": "",
            "equalizer_fit_mode": "",
            "equalizer_normalization_applied": "",
            "equalizer_final_is_normalized": "",
            "equalizer_final_sum": "",
            "equalizer_fit_residual": "",
            "equalizer_valid_coverage": "",
            "equalizer_calibration_valid": str(bool(calibration_valid)),
            "equalizer_frame_sha256": (
                frame_rgb_sha256(rgb) if rgb is not None else ""
            ),
            "equalizer_frame_sha256_algorithm": (
                "rgb-array-v1" if rgb is not None else ""
            ),
        }
        for family in ("raw", "final", "normalized"):
            for suffix in EQUALIZER_CSV_SUFFIX.values():
                columns[f"equalizer_{family}_{suffix}"] = ""
            columns[f"equalizer_{family}_HTR"] = ""
        if payload is None:
            return columns

        canonical = validate_equalizer_payload(payload)
        columns.update({
            "equalizer_schema_version": canonical["schema_version"],
            "equalizer_source": canonical["label_source"],
            "equalizer_confidence": canonical["confidence"],
            "equalizer_argmax": canonical["argmax"],
            "equalizer_fit_mode": canonical["fit_mode"],
            "equalizer_normalization_applied": str(
                bool(canonical["normalization_applied"])
            ),
            "equalizer_final_is_normalized": str(
                bool(canonical["final_is_normalized"])
            ),
            "equalizer_final_sum": f"{float(canonical['final_sum']):.6f}",
            "equalizer_fit_residual": f"{float(canonical['residual_rms']):.6f}",
            "equalizer_valid_coverage": f"{float(canonical['valid_coverage']):.6f}",
        })
        for family in ("raw", "final", "normalized"):
            values = canonical[f"{family}_weights"]
            for label, suffix in EQUALIZER_CSV_SUFFIX.items():
                value = values[label]
                columns[f"equalizer_{family}_{suffix}"] = (
                    "" if value is None else f"{float(value):.8f}"
                )
            # HTR has no canonical basis and remains a CSV null.
            columns[f"equalizer_{family}_HTR"] = ""
        return columns

    @classmethod
    def _human_primary_columns(
        cls,
        *,
        reconstruction: str,
        label_source: str,
        labeler: str,
        confidence: object,
        blind_to_model: bool,
        blind_to_equalizer: bool,
        frame_path: Path,
        capture_metadata: Mapping[str, object],
        session_dir: Path,
    ) -> dict[str, object]:
        """Validate a human judgment and bind it to one exact decoded frame."""
        allowed = {
            "none/weak", "Twinned (2x1)", "c(6x2)", "rt13xrt13",
            "HTR", "unknown", "artifact",
        }
        if reconstruction not in allowed:
            raise ValueError(f"Unsupported human primary label {reconstruction!r}")
        if label_source not in {"human_blind_primary", "human_assisted_primary"}:
            raise ValueError("Human primary label_source is invalid")
        if label_source == "human_blind_primary" and not (
            blind_to_model and blind_to_equalizer
        ):
            raise ValueError("Blind primary source requires both blindness flags")
        labeler_text = str(labeler or "").strip()
        if not labeler_text:
            raise ValueError("Human primary labeler is required")

        confidence_text = str(confidence or "").strip()
        if confidence_text:
            confidence_value = float(confidence_text)
            if not math.isfinite(confidence_value) or not 0.0 <= confidence_value <= 1.0:
                raise ValueError("Human primary confidence must be in [0, 1]")
            confidence_text = f"{confidence_value:.3f}"

        exact_path = cls._resolve_existing_event_frame(session_dir, str(frame_path))
        metadata = dict(capture_metadata)
        required = (
            "capture_backend", "captured_at_utc", "capture_sequence",
            "view_segment_id", "visual_history_generation", "session_id",
            "gun_aligned", "realignment_active",
        )
        missing = [name for name in required if metadata.get(name) in (None, "")]
        if missing:
            raise ValueError(
                "Human primary capture binding is missing: " + ", ".join(missing)
            )
        sequence = int(metadata["capture_sequence"])
        segment = int(metadata["view_segment_id"])
        generation = int(metadata["visual_history_generation"])
        if sequence <= 0 or segment < 0 or generation < 0:
            raise ValueError("Human primary capture identity is invalid")
        if metadata["gun_aligned"] is not True or metadata["realignment_active"] is not False:
            raise ValueError("Human primary frame is not from a stable aligned view")
        normalize_utc_timestamp(
            str(metadata["captured_at_utc"]), field_name="human captured_at_utc",
        )
        try:
            from PIL import Image

            with Image.open(exact_path) as image:
                rgb = np.asarray(image.convert("RGB"), dtype=np.uint8)
        except Exception as exc:
            raise ValueError("Human primary frame cannot be decoded") from exc

        return {
            "human_primary_reconstruction": reconstruction,
            "human_label_source": label_source,
            "human_labeler": labeler_text,
            "human_confidence": confidence_text,
            "human_blind_to_model": str(bool(blind_to_model)),
            "human_blind_to_equalizer": str(bool(blind_to_equalizer)),
            "human_primary_frame_path": str(exact_path),
            "human_primary_frame_sha256": frame_rgb_sha256(rgb),
            "human_primary_frame_sha256_algorithm": "rgb-array-v1",
            "human_primary_capture_backend": str(metadata["capture_backend"]),
            "human_primary_captured_at_utc": str(metadata["captured_at_utc"]),
            "human_primary_capture_sequence": sequence,
            "human_primary_view_segment_id": segment,
            "human_primary_visual_history_generation": generation,
            "human_primary_session_id": str(metadata["session_id"]),
        }

    @staticmethod
    def _utc_now_text() -> str:
        return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")

    def _human_labeling_state_path(self) -> Optional[Path]:
        if self._session_dir is None:
            return None
        return self._session_dir / self.HUMAN_LABELING_STATE_FILE

    def _new_human_labeling_state(self) -> dict[str, object]:
        if self._session_dir is None:
            raise RuntimeError("No growth session is available")
        now = self._utc_now_text()
        return {
            "schema_version": self.HUMAN_LABELING_STATE_SCHEMA_VERSION,
            "session_id": self._session_dir.name,
            "created_at_utc": now,
            "last_updated_at_utc": now,
            "created_runtime_id": self._human_labeling_runtime_id,
            "active_runtime_id": self._human_labeling_runtime_id,
            "attach_count": 0,
            "restart_count": 0,
            "blind_mode_active": False,
            "blind_mode_labeler": "",
            "blind_mode_entered_at_utc": "",
            "global_classifier_revealed": False,
            "global_equalizer_revealed": False,
            "revealed_frames": {},
        }

    def _validate_human_labeling_state(
        self, raw: object,
    ) -> dict[str, object]:
        """Validate the persisted proof state; malformed state fails closed."""
        if not isinstance(raw, dict):
            raise ValueError("Human-labeling state must be a JSON object")
        required = {
            "schema_version", "session_id", "created_at_utc",
            "last_updated_at_utc", "created_runtime_id",
            "active_runtime_id", "attach_count", "restart_count",
            "blind_mode_active", "blind_mode_labeler",
            "blind_mode_entered_at_utc", "global_classifier_revealed",
            "global_equalizer_revealed", "revealed_frames",
        }
        missing = sorted(required - set(raw))
        if missing:
            raise ValueError(
                "Human-labeling state is missing: " + ", ".join(missing)
            )
        if raw.get("schema_version") != self.HUMAN_LABELING_STATE_SCHEMA_VERSION:
            raise ValueError("Human-labeling state schema is unsupported")
        if self._session_dir is None or raw.get("session_id") != self._session_dir.name:
            raise ValueError("Human-labeling state belongs to another session")
        for field in (
            "blind_mode_active", "global_classifier_revealed",
            "global_equalizer_revealed",
        ):
            if not isinstance(raw.get(field), bool):
                raise ValueError(f"Human-labeling state {field} must be boolean")
        for field in ("attach_count", "restart_count"):
            value = raw.get(field)
            if not isinstance(value, int) or value < 0:
                raise ValueError(f"Human-labeling state {field} is invalid")
        revealed = raw.get("revealed_frames")
        if not isinstance(revealed, dict):
            raise ValueError("Human-labeling revealed_frames must be an object")
        for digest, exposures in revealed.items():
            if (
                not isinstance(digest, str)
                or len(digest) != 64
                or not isinstance(exposures, dict)
                or any(kind not in {"classifier", "equalizer"} for kind in exposures)
            ):
                raise ValueError("Human-labeling exposure record is invalid")
        return dict(raw)

    def _read_human_labeling_state(self) -> Optional[dict[str, object]]:
        path = self._human_labeling_state_path()
        if path is None or not path.is_file():
            self._human_labeling_state_trusted = False
            return None
        try:
            state = self._validate_human_labeling_state(
                json.loads(path.read_text(encoding="utf-8"))
            )
        except (OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
            self._human_labeling_state_trusted = False
            log.error("Human-labeling proof state is untrusted: %s", exc)
            return None
        self._human_labeling_state_trusted = True
        return state

    def _write_human_labeling_state(self, state: Mapping[str, object]) -> bool:
        path = self._human_labeling_state_path()
        if path is None:
            return False
        candidate = dict(state)
        candidate["last_updated_at_utc"] = self._utc_now_text()
        try:
            validated = self._validate_human_labeling_state(candidate)
            payload = (
                json.dumps(validated, sort_keys=True, indent=2) + "\n"
            ).encode("utf-8")
        except (TypeError, ValueError) as exc:
            log.error("Refusing invalid human-labeling proof state: %s", exc)
            self._human_labeling_state_trusted = False
            return False
        written = self._atomic_write_bytes(path, payload)
        self._human_labeling_state_trusted = written
        return written

    def _initialize_human_labeling_state(self) -> bool:
        """Create proof state only for a session born in this GUI runtime."""
        return self._write_human_labeling_state(self._new_human_labeling_state())

    def begin_human_labeling_review(self, labeler: str = "") -> dict[str, object]:
        """Attach the labeling UI and force an explicit new blind-mode entry.

        A persisted ``blind_mode_active`` value is never restored implicitly.
        The runtime ID change records a restart; until the operator explicitly
        enters blind mode again, subsequent judgments are assisted.
        """
        state = self._read_human_labeling_state()
        if state is None:
            return {
                "audit_trusted": False,
                "blind_mode_active": False,
                "gold_eligible": False,
                "reason": "No trusted labeling-state record exists for this session.",
            }
        previous_runtime = str(state.get("active_runtime_id", ""))
        if previous_runtime and previous_runtime != self._human_labeling_runtime_id:
            state["restart_count"] = int(state["restart_count"]) + 1
        state["active_runtime_id"] = self._human_labeling_runtime_id
        state["attach_count"] = int(state["attach_count"]) + 1
        state["blind_mode_active"] = False
        state["blind_mode_labeler"] = str(labeler or "").strip()
        state["blind_mode_entered_at_utc"] = ""
        if not self._write_human_labeling_state(state):
            return {
                "audit_trusted": False,
                "blind_mode_active": False,
                "gold_eligible": False,
                "reason": "The labeling-state record could not be persisted.",
            }
        return {
            "audit_trusted": True,
            "blind_mode_active": False,
            "gold_eligible": False,
            "reason": "Enter blind labeling mode before assigning gold labels.",
        }

    def set_human_blind_labeling_mode(
        self, enabled: bool, *, labeler: str,
    ) -> dict[str, object]:
        """Persist an explicit blind-mode transition and report eligibility."""
        labeler_text = str(labeler or "").strip()
        if enabled and not labeler_text:
            raise ValueError("A labeler name is required for blind labeling mode")
        state = self._read_human_labeling_state()
        if state is None:
            raise ValueError(
                "Blind labeling cannot be audited for this session; "
                "labels will be assisted only"
            )
        if state.get("active_runtime_id") != self._human_labeling_runtime_id:
            raise ValueError(
                "The labeling UI was restarted; attach it again before entering blind mode"
            )
        state["blind_mode_active"] = bool(enabled)
        state["blind_mode_labeler"] = labeler_text if enabled else ""
        state["blind_mode_entered_at_utc"] = (
            self._utc_now_text() if enabled else ""
        )
        if not self._write_human_labeling_state(state):
            raise OSError("Blind labeling state could not be persisted")
        contaminated = bool(
            state["global_classifier_revealed"]
            or state["global_equalizer_revealed"]
        )
        return {
            "audit_trusted": True,
            "blind_mode_active": bool(enabled),
            "gold_eligible": bool(enabled) and not contaminated,
            "reason": (
                "Classifier or Equalizer output was already displayed; "
                "new labels remain assisted."
                if enabled and contaminated
                else "Blind labeling mode is active."
                if enabled
                else "Blind labeling mode is off."
            ),
        }

    def record_human_label_output_exposure(
        self, kind: str, *, frame_path: Optional[str | Path] = None,
    ) -> bool:
        """Persist that classifier/Equalizer output became visible.

        The global flag is deliberately irreversible for the session.  This
        fail-closed rule avoids claiming an unbiased label after the grower
        has already seen a machine-generated answer.
        """
        if kind not in {"classifier", "equalizer"}:
            raise ValueError("Exposure kind must be classifier or equalizer")
        state = self._read_human_labeling_state()
        if state is None:
            # With no trusted proof state every future annotation is already
            # forced to assisted. Showing an aid cannot create false gold.
            return True
        state[f"global_{kind}_revealed"] = True
        if frame_path is not None:
            try:
                exact_path = self._resolve_existing_event_frame(
                    self._session_dir, frame_path,
                )
                from PIL import Image

                with Image.open(exact_path) as image:
                    digest = frame_rgb_sha256(
                        np.asarray(image.convert("RGB"), dtype=np.uint8)
                    )
                revealed = dict(state.get("revealed_frames", {}))
                exposures = dict(revealed.get(digest, {}))
                exposures[kind] = self._utc_now_text()
                revealed[digest] = exposures
                state["revealed_frames"] = revealed
            except (OSError, TypeError, ValueError):
                # The global exposure is still auditable even if an old frame
                # disappeared before it could be hashed.
                pass
        state["blind_mode_active"] = False
        state["blind_mode_labeler"] = ""
        state["blind_mode_entered_at_utc"] = ""
        return self._write_human_labeling_state(state)

    def human_primary_provenance(
        self, *, frame_path: str | Path, labeler: str,
    ) -> dict[str, object]:
        """Derive source/blind flags solely from persisted proof state."""
        if self._session_dir is None:
            raise ValueError("No session is attached")
        labeler_text = str(labeler or "").strip()
        if not labeler_text:
            raise ValueError("Human primary labeler is required")
        exact_path = self._resolve_existing_event_frame(
            self._session_dir, frame_path,
        )
        from PIL import Image

        with Image.open(exact_path) as image:
            digest = frame_rgb_sha256(
                np.asarray(image.convert("RGB"), dtype=np.uint8)
            )
        state = self._read_human_labeling_state()
        if state is None:
            return {
                "human_label_source": "human_assisted_primary",
                "human_blind_to_model": False,
                "human_blind_to_equalizer": False,
                "frame_sha256": digest,
                "reason": "Blind provenance is unavailable or untrusted.",
            }
        exposures = dict(state.get("revealed_frames", {})).get(digest, {})
        blind_to_model = not bool(
            state["global_classifier_revealed"]
            or (isinstance(exposures, dict) and exposures.get("classifier"))
        )
        blind_to_equalizer = not bool(
            state["global_equalizer_revealed"]
            or (isinstance(exposures, dict) and exposures.get("equalizer"))
        )
        explicit_mode = bool(
            state["blind_mode_active"]
            and state.get("active_runtime_id") == self._human_labeling_runtime_id
            and state.get("blind_mode_labeler") == labeler_text
        )
        gold = explicit_mode and blind_to_model and blind_to_equalizer
        return {
            "human_label_source": (
                "human_blind_primary" if gold else "human_assisted_primary"
            ),
            "human_blind_to_model": blind_to_model,
            "human_blind_to_equalizer": blind_to_equalizer,
            "frame_sha256": digest,
            "reason": (
                "Persisted blind-mode proof is valid."
                if gold
                else "No valid uncontaminated blind-mode proof exists."
            ),
        }

    def _append_human_primary_label(
        self,
        *,
        human_columns: Mapping[str, object],
        capture_metadata: Mapping[str, object],
        annotation_id: str = "",
        qc_label: str = "",
        timestamp: str = "",
    ) -> dict[str, str]:
        """Append one immutable human judgment and fsync it before summary."""
        if self._session_dir is None:
            raise ValueError("No session is attached")
        path = self._session_dir / "human_primary_labels.csv"
        rows: list[dict[str, str]] = []
        if path.exists() and path.stat().st_size:
            try:
                with open(path, newline="", encoding="utf-8") as stream:
                    reader = csv.DictReader(stream)
                    if tuple(reader.fieldnames or ()) != tuple(
                        self.HUMAN_PRIMARY_LABEL_FIELDS
                    ):
                        raise ValueError(
                            "human_primary_labels.csv schema is not trusted"
                        )
                    rows = [dict(row) for row in reader]
            except OSError as exc:
                raise OSError("Human primary audit log could not be read") from exc
        indices: list[int] = []
        annotation_ids: set[str] = set()
        for row in rows:
            try:
                idx = int(row.get("label_idx", ""))
            except (TypeError, ValueError) as exc:
                raise ValueError("Human primary audit log has an invalid label_idx") from exc
            identifier = str(row.get("annotation_id", ""))
            if idx <= 0 or not identifier or identifier in annotation_ids:
                raise ValueError("Human primary audit log identity is invalid")
            indices.append(idx)
            annotation_ids.add(identifier)
        if len(indices) != len(set(indices)):
            raise ValueError("Human primary audit log has duplicate label_idx values")

        identifier = str(annotation_id or uuid.uuid4().hex).strip()
        if not identifier or identifier in annotation_ids:
            raise ValueError("Human primary annotation_id must be unique")
        label_idx = max(indices, default=0) + 1
        exact_path = Path(str(human_columns["human_primary_frame_path"]))
        try:
            portable_path = exact_path.relative_to(self._session_dir).as_posix()
        except ValueError:
            portable_path = str(exact_path)
        session_id = str(human_columns["human_primary_session_id"])
        row = {
            "schema_version": "1",
            "annotation_id": identifier,
            "label_idx": str(label_idx),
            "label_timestamp_iso": timestamp or self._utc_now_text(),
            "session_id": session_id,
            "acquisition_run": session_id,
            "frame_path": portable_path,
            "frame_sha256": str(human_columns["human_primary_frame_sha256"]),
            "frame_sha256_algorithm": str(
                human_columns["human_primary_frame_sha256_algorithm"]
            ),
            "capture_sequence": str(
                human_columns["human_primary_capture_sequence"]
            ),
            "view_segment_id": str(
                human_columns["human_primary_view_segment_id"]
            ),
            "capture_backend": str(
                human_columns["human_primary_capture_backend"]
            ),
            "captured_at_utc": str(
                human_columns["human_primary_captured_at_utc"]
            ),
            "gun_aligned": str(bool(capture_metadata["gun_aligned"])),
            "realignment_active": str(bool(
                capture_metadata["realignment_active"]
            )),
            "human_primary_reconstruction": str(
                human_columns["human_primary_reconstruction"]
            ),
            "human_label_source": str(human_columns["human_label_source"]),
            "human_labeler": str(human_columns["human_labeler"]),
            "human_confidence": str(human_columns["human_confidence"]),
            "human_blind_to_model": str(
                human_columns["human_blind_to_model"]
            ),
            "human_blind_to_equalizer": str(
                human_columns["human_blind_to_equalizer"]
            ),
            "qc_label": str(qc_label or capture_metadata.get("qc_label", "")),
        }
        try:
            needs_header = not path.exists() or path.stat().st_size == 0
            with open(path, "a", newline="", encoding="utf-8") as stream:
                writer = csv.DictWriter(
                    stream, fieldnames=self.HUMAN_PRIMARY_LABEL_FIELDS,
                )
                if needs_header:
                    writer.writeheader()
                writer.writerow(row)
                stream.flush()
                os.fsync(stream.fileno())
        except OSError as exc:
            raise OSError("Human primary audit append failed") from exc
        return row

    def read_human_primary_labels(self) -> list[dict[str, str]]:
        """Read the append-only human-primary audit log without summarizing."""
        if self._session_dir is None:
            return []
        path = self._session_dir / "human_primary_labels.csv"
        if not path.is_file():
            return []
        try:
            with open(path, newline="", encoding="utf-8") as stream:
                reader = csv.DictReader(stream)
                if tuple(reader.fieldnames or ()) != tuple(
                    self.HUMAN_PRIMARY_LABEL_FIELDS
                ):
                    return []
                return [dict(row) for row in reader]
        except OSError:
            return []

    @classmethod
    def _snapshot_capture_columns(cls, snapshot: RheedFrameSnapshot) -> dict:
        age_ms = snapshot.logging_age_ms()
        if not math.isfinite(age_ms):
            raise ValueError("RHEED snapshot has no valid monotonic receive time")
        return cls._capture_columns({
            "capture_backend": snapshot.capture_backend,
            "captured_at_utc": snapshot.captured_at_utc,
            "capture_sequence": snapshot.capture_sequence,
            "frame_age_ms": age_ms,
            "source_hwnd": snapshot.source_hwnd,
            "capture_geometry_id": snapshot.capture_geometry_id,
        })

    @staticmethod
    def _validate_capture_identity(
        *,
        captured_at_utc: str,
        capture_sequence: int,
        received_monotonic_ns: int,
        capture_backend: str,
        source_hwnd: int,
        capture_geometry_id: str,
        camera_width: int,
        camera_height: int,
    ) -> None:
        """Reject incomplete provenance before a calibrated write."""
        normalize_utc_timestamp(
            captured_at_utc,
            field_name="RHEED captured_at_utc",
        )
        if int(capture_sequence) <= 0:
            raise ValueError("RHEED capture_sequence must be positive")
        if int(received_monotonic_ns) <= 0:
            raise ValueError("RHEED monotonic receive timestamp is missing")
        backend = str(capture_backend).strip()
        if not backend:
            raise ValueError("RHEED capture backend is missing")
        if not str(capture_geometry_id).strip():
            raise ValueError("RHEED capture geometry identity is missing")
        if int(camera_width) <= 0 or int(camera_height) <= 0:
            raise ValueError("RHEED camera dimensions must be positive")
        if backend == "wgc" and int(source_hwnd) <= 0:
            raise ValueError("WGC source HWND must be positive")

    @staticmethod
    def _validate_equalizer_context(
        calibration: CalibrationRecord,
        snapshot: RheedFrameSnapshot,
    ) -> None:
        """Reject stale or untraceable calibration/frame combinations."""
        if not isinstance(calibration, CalibrationRecord):
            raise TypeError("calibration must be a CalibrationRecord")
        if not isinstance(snapshot, RheedFrameSnapshot):
            raise TypeError("snapshot must be a RheedFrameSnapshot")
        if not calibration.grower_accepted or calibration.invalidated_reason:
            raise ValueError("Equalizer calibration is not accepted and active")
        if not calibration.calibration_id or not calibration.basis_bundle_id:
            raise ValueError("Equalizer calibration provenance is incomplete")
        if not calibration.gun_aligned or calibration.realignment_active:
            raise ValueError("Calibration was not accepted on a stable aligned view")
        if not snapshot.gun_aligned or snapshot.realignment_active:
            raise ValueError("Snapshot is not from a stable aligned view")
        if not calibration.session_id or not snapshot.session_id:
            raise ValueError("Equalizer session identity is missing")
        if calibration.view_segment_id is None or snapshot.view_segment_id is None:
            raise ValueError("RHEED view segment is missing")
        GrowthLogger._validate_capture_identity(
            captured_at_utc=calibration.captured_at_utc,
            capture_sequence=calibration.capture_sequence,
            received_monotonic_ns=calibration.received_monotonic_ns,
            capture_backend=calibration.capture_backend,
            source_hwnd=calibration.source_hwnd,
            capture_geometry_id=calibration.capture_geometry_id,
            camera_width=calibration.camera_width,
            camera_height=calibration.camera_height,
        )
        GrowthLogger._validate_capture_identity(
            captured_at_utc=snapshot.captured_at_utc,
            capture_sequence=snapshot.capture_sequence,
            received_monotonic_ns=snapshot.received_monotonic_ns,
            capture_backend=snapshot.capture_backend,
            source_hwnd=snapshot.source_hwnd,
            capture_geometry_id=snapshot.capture_geometry_id,
            camera_width=snapshot.camera_width,
            camera_height=snapshot.camera_height,
        )

        comparisons = (
            ("session", calibration.session_id, snapshot.session_id),
            ("view segment", calibration.view_segment_id, snapshot.view_segment_id),
            (
                "visual-history generation",
                calibration.visual_history_generation,
                snapshot.visual_history_generation,
            ),
            ("capture backend", calibration.capture_backend, snapshot.capture_backend),
            ("source HWND", calibration.source_hwnd, snapshot.source_hwnd),
            (
                "capture geometry",
                calibration.capture_geometry_id,
                snapshot.capture_geometry_id,
            ),
            ("camera width", calibration.camera_width, snapshot.camera_width),
            ("camera height", calibration.camera_height, snapshot.camera_height),
        )
        for label, expected, actual in comparisons:
            if expected != actual:
                raise ValueError(
                    f"Snapshot {label} {actual!r} does not match "
                    f"calibration {expected!r}"
                )
        if snapshot.received_monotonic_ns <= 0:
            raise ValueError("RHEED snapshot receive timestamp is missing")

    @staticmethod
    def _resolve_existing_event_frame(
        session_dir: Path,
        frame_path: str | Path,
    ) -> Path:
        """Resolve an already captured event frame inside session/frames."""
        value = str(frame_path).strip()
        if not value:
            raise ValueError("Equalizer event frame_path cannot be blank")
        session_dir = Path(session_dir)
        candidate = Path(value)
        if not candidate.is_absolute():
            candidate = session_dir / candidate
        try:
            resolved = candidate.resolve(strict=True)
            frames_root = (session_dir / "frames").resolve(strict=True)
            resolved.relative_to(frames_root)
        except (OSError, RuntimeError, ValueError) as exc:
            raise ValueError(
                "Equalizer event frame must exist inside the session frames directory"
            ) from exc
        if not resolved.is_file():
            raise ValueError("Equalizer event frame_path is not a file")
        return resolved

    @staticmethod
    def _save_rgb_bmp(frame: np.ndarray, path: Path) -> bool:
        """Atomically write one RGB BMP, reporting all encoder failures."""
        destination = Path(path)
        descriptor = -1
        temporary: Optional[Path] = None
        errors: list[str] = []
        try:
            destination.parent.mkdir(parents=True, exist_ok=True)
            descriptor, name = tempfile.mkstemp(
                prefix=f".{destination.stem}.",
                suffix=".tmp.bmp",
                dir=str(destination.parent),
            )
            os.close(descriptor)
            descriptor = -1
            temporary = Path(name)

            encoded = False
            try:
                from PIL import Image

                Image.fromarray(frame).save(str(temporary), format="BMP")
                encoded = True
            except Exception as exc:  # optional encoder boundary
                errors.append(f"PIL: {exc}")

            if not encoded:
                try:
                    import cv2

                    encoded = bool(cv2.imwrite(
                        str(temporary),
                        cv2.cvtColor(frame, cv2.COLOR_RGB2BGR),
                    ))
                    if not encoded:
                        errors.append("OpenCV: imwrite returned False")
                except Exception as exc:  # optional encoder boundary
                    errors.append(f"OpenCV: {exc}")

            if not encoded or not temporary.exists() or temporary.stat().st_size <= 0:
                raise OSError("; ".join(errors) or "encoder produced no image")
            # Windows' CRT rejects fsync on a read-only descriptor.
            with open(temporary, "r+b") as stream:
                os.fsync(stream.fileno())
            os.replace(temporary, destination)
            temporary = None
            return True
        except (OSError, TypeError, ValueError) as exc:
            detail = "; ".join(errors)
            if detail and detail not in str(exc):
                detail = f"{exc}; {detail}"
            else:
                detail = str(exc)
            log.error("Failed to save RHEED image %s: %s", destination, detail)
            return False
        finally:
            if descriptor >= 0:
                try:
                    os.close(descriptor)
                except OSError:
                    pass
            if temporary is not None:
                try:
                    temporary.unlink(missing_ok=True)
                except OSError:
                    pass

    @staticmethod
    def _atomic_write_bytes(path: Path, payload: bytes) -> bool:
        """Write bytes through a same-directory temp file and durable replace."""
        destination = Path(path)
        descriptor = -1
        temporary: Optional[Path] = None
        try:
            destination.parent.mkdir(parents=True, exist_ok=True)
            descriptor, name = tempfile.mkstemp(
                prefix=f".{destination.name}.",
                suffix=".tmp",
                dir=str(destination.parent),
            )
            temporary = Path(name)
            with os.fdopen(descriptor, "wb") as stream:
                descriptor = -1
                stream.write(payload)
                stream.flush()
                os.fsync(stream.fileno())
            os.replace(temporary, destination)
            temporary = None
            return True
        except (OSError, TypeError, ValueError) as exc:
            log.error("Failed to atomically write %s: %s", destination, exc)
            return False
        finally:
            if descriptor >= 0:
                try:
                    os.close(descriptor)
                except OSError:
                    pass
            if temporary is not None:
                try:
                    temporary.unlink(missing_ok=True)
                except OSError:
                    pass

    @staticmethod
    def _atomic_write_csv(
        path: Path,
        fieldnames: Iterable[str],
        rows: Iterable[Mapping[str, object]],
    ) -> bool:
        """Durably replace a CSV without exposing a truncated destination."""
        destination = Path(path)
        descriptor = -1
        temporary: Optional[Path] = None
        try:
            destination.parent.mkdir(parents=True, exist_ok=True)
            descriptor, name = tempfile.mkstemp(
                prefix=f".{destination.name}.",
                suffix=".tmp",
                dir=str(destination.parent),
            )
            temporary = Path(name)
            with os.fdopen(
                descriptor, "w", newline="", encoding="utf-8",
            ) as stream:
                descriptor = -1
                writer = csv.DictWriter(stream, fieldnames=tuple(fieldnames))
                writer.writeheader()
                writer.writerows(rows)
                stream.flush()
                os.fsync(stream.fileno())
            os.replace(temporary, destination)
            temporary = None
            return True
        except (OSError, TypeError, ValueError) as exc:
            log.error("Failed to atomically replace CSV %s: %s", destination, exc)
            return False
        finally:
            if descriptor >= 0:
                try:
                    os.close(descriptor)
                except OSError:
                    pass
            if temporary is not None:
                try:
                    temporary.unlink(missing_ok=True)
                except OSError:
                    pass

    @classmethod
    def _live_label_transaction_path(cls, session_dir: Path) -> Path:
        return Path(session_dir) / cls.LIVE_LABEL_TRANSACTION_FILE

    @classmethod
    def _calibration_transaction_path(cls, session_dir: Path) -> Path:
        return Path(session_dir) / cls.CALIBRATION_TRANSACTION_FILE

    @classmethod
    def _auto_capture_transaction_path(cls, session_dir: Path) -> Path:
        return Path(session_dir) / cls.AUTO_CAPTURE_TRANSACTION_FILE

    @staticmethod
    def _resolve_session_relative_path(session_dir: Path, value: object) -> Path:
        """Resolve one WAL-owned path without permitting directory escape."""
        relative = Path(str(value).replace("\\", "/"))
        if relative.is_absolute() or not relative.parts or ".." in relative.parts:
            raise ValueError("Live-label transaction contains an unsafe path")
        return Path(session_dir) / relative

    @classmethod
    def _recover_live_label_transaction(cls, session_dir: Path) -> bool:
        """Finish or roll back one interrupted Live-label transaction.

        The CSV replacement is atomic, so presence of the exact WAL row is the
        commit decision.  A committed row keeps/finalizes its image; an absent
        row removes both hidden and final image paths.  Malformed or conflicting
        state is retained and rejected so evidence is never silently guessed.
        """
        session_dir = Path(session_dir)
        marker = cls._live_label_transaction_path(session_dir)
        if not marker.exists():
            return True
        try:
            payload = json.loads(marker.read_text(encoding="utf-8"))
            if (
                not isinstance(payload, dict)
                or payload.get("schema_version")
                != cls.LIVE_LABEL_TRANSACTION_SCHEMA_VERSION
            ):
                raise ValueError("unsupported live-label WAL schema")
            row = payload.get("row")
            if not isinstance(row, dict) or set(row) != set(cls.LIVE_LABEL_FIELDS):
                raise ValueError("live-label WAL row does not match the CSV schema")
            pending_path = cls._resolve_session_relative_path(
                session_dir, payload.get("pending_frame_path", ""),
            )
            final_path = cls._resolve_session_relative_path(
                session_dir, payload.get("final_frame_path", ""),
            )
            if str(row.get("frame_path", "")) != str(final_path):
                raise ValueError("live-label WAL frame path does not match its row")

            csv_path = session_dir / "live_labels.csv"
            with open(csv_path, newline="", encoding="utf-8") as stream:
                reader = csv.DictReader(stream)
                if tuple(reader.fieldnames or ()) != tuple(cls.LIVE_LABEL_FIELDS):
                    raise ValueError("live-label CSV schema does not match the WAL")
                rows = list(reader)

            expected = {
                field: "" if row.get(field) is None else str(row.get(field))
                for field in cls.LIVE_LABEL_FIELDS
            }
            same_index = [
                existing for existing in rows
                if existing.get("label_idx") == expected["label_idx"]
            ]
            if len(same_index) > 1:
                raise ValueError("live-label CSV contains a duplicate label index")
            committed = bool(same_index)
            if committed and same_index[0] != expected:
                raise ValueError("live-label WAL conflicts with the committed CSV row")

            if committed:
                if final_path.exists():
                    if pending_path.exists():
                        if (
                            hashlib.sha256(pending_path.read_bytes()).digest()
                            != hashlib.sha256(final_path.read_bytes()).digest()
                        ):
                            raise ValueError(
                                "live-label pending and final images conflict"
                            )
                        pending_path.unlink()
                elif pending_path.exists():
                    os.replace(pending_path, final_path)
                else:
                    raise ValueError(
                        "committed live-label row has no recoverable image"
                    )
                if final_path.stat().st_size <= 0:
                    raise ValueError("recovered live-label image is empty")
            else:
                # The atomic CSV commit did not happen.  Any image named in
                # this WAL belongs solely to the abandoned transaction.
                pending_path.unlink(missing_ok=True)
                final_path.unlink(missing_ok=True)

            # A hard stop can bypass the atomic helpers' ``finally`` blocks.
            # Their uniquely named same-directory temporaries remain owned by
            # this WAL and must not become silent, untracked image/CSV debris.
            temporary_paths = (
                *pending_path.parent.glob(
                    f".{pending_path.stem}.*.tmp.bmp"
                ),
                *csv_path.parent.glob(f".{csv_path.name}.*.tmp"),
            )
            for temporary_path in temporary_paths:
                temporary_path.unlink(missing_ok=True)

            marker.unlink()
            return True
        except (OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
            log.error(
                "Live-label transaction recovery failed for %s: %s",
                marker,
                exc,
            )
            return False

    @classmethod
    def _recover_calibration_transaction(cls, session_dir: Path) -> bool:
        """Resolve a calibration evidence/journal transaction after restart.

        The append-only JSONL line is the commit decision.  If the exact line
        is present, its evidence must also be present and hash-valid.  If it
        is absent, the WAL-owned evidence is rolled back.  Any ambiguity is
        retained fail-closed for operator inspection.
        """
        session_dir = Path(session_dir)
        marker = cls._calibration_transaction_path(session_dir)
        if not marker.exists():
            return True
        try:
            payload = json.loads(marker.read_text(encoding="utf-8"))
            if (
                not isinstance(payload, dict)
                or payload.get("schema_version")
                != cls.CALIBRATION_TRANSACTION_SCHEMA_VERSION
            ):
                raise ValueError("unsupported calibration WAL schema")
            event = payload.get("event")
            if not isinstance(event, dict):
                raise ValueError("calibration WAL event is missing")
            if (
                event.get("journal_schema_version") != 1
                or event.get("event") != "accepted"
            ):
                raise ValueError("calibration WAL does not contain acceptance")
            calibration_payload = event.get("calibration")
            if not isinstance(calibration_payload, dict):
                raise ValueError("calibration WAL record is missing")
            calibration_id = str(payload.get("calibration_id", ""))
            if (
                not calibration_id
                or str(calibration_payload.get("calibration_id", ""))
                != calibration_id
            ):
                raise ValueError("calibration WAL identity is inconsistent")
            evidence_relative = str(payload.get("evidence_path", ""))
            if (
                evidence_relative
                != str(calibration_payload.get("orientation_evidence_path", ""))
                or evidence_relative
                != str(event.get("orientation_evidence_path", ""))
            ):
                raise ValueError("calibration WAL evidence path is inconsistent")
            evidence_path = cls._resolve_session_relative_path(
                session_dir, evidence_relative,
            )
            evidence_hash = str(payload.get("evidence_sha256", "")).lower()
            if (
                len(evidence_hash) != 64
                or any(character not in "0123456789abcdef" for character in evidence_hash)
                or evidence_hash
                != str(calibration_payload.get(
                    "orientation_evidence_sha256", "",
                )).lower()
            ):
                raise ValueError("calibration WAL evidence hash is invalid")

            journal_path = session_dir / "equalizer_calibrations.jsonl"
            untrusted_path = journal_path.with_suffix(".untrusted")
            if untrusted_path.exists():
                raise ValueError("calibration journal is marked untrusted")

            matching_events: list[dict] = []
            conflicting_identity = False
            if journal_path.exists():
                with open(journal_path, "r", encoding="utf-8") as stream:
                    for raw_line in stream:
                        if not raw_line.strip():
                            continue
                        journal_event = json.loads(raw_line)
                        if not isinstance(journal_event, dict):
                            raise ValueError("calibration journal line is not an object")
                        journal_calibration = journal_event.get("calibration")
                        journal_id = (
                            str(journal_calibration.get("calibration_id", ""))
                            if isinstance(journal_calibration, dict)
                            else ""
                        )
                        if journal_event == event:
                            matching_events.append(journal_event)
                        elif journal_id == calibration_id:
                            conflicting_identity = True
            if len(matching_events) > 1 or conflicting_identity:
                raise ValueError("calibration WAL conflicts with the journal")

            if matching_events:
                if not evidence_path.exists():
                    raise ValueError("committed calibration has no evidence image")
                digest = hashlib.sha256(evidence_path.read_bytes()).hexdigest()
                if digest != evidence_hash:
                    raise ValueError("committed calibration evidence hash mismatch")
            elif evidence_path.exists():
                digest = hashlib.sha256(evidence_path.read_bytes()).hexdigest()
                if digest != evidence_hash:
                    raise ValueError("uncommitted calibration evidence hash mismatch")
                evidence_path.unlink()

            for temporary_path in evidence_path.parent.glob(
                f".{evidence_path.name}.*.tmp"
            ):
                temporary_path.unlink(missing_ok=True)
            marker.unlink()
            return True
        except (OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
            log.error(
                "Calibration transaction recovery failed for %s: %s",
                marker,
                exc,
            )
            return False

    @staticmethod
    def _remove_auto_capture_directory(path: Path) -> bool:
        """Remove one WAL-owned flat event directory without guessing."""
        directory = Path(path)
        if not directory.exists():
            return True
        try:
            if directory.is_symlink() or not directory.is_dir():
                raise ValueError("auto-capture transaction path is not a directory")
            children = tuple(directory.iterdir())
            if any(child.is_dir() and not child.is_symlink() for child in children):
                raise ValueError("auto-capture transaction directory is not flat")
            for child in children:
                child.unlink()
            directory.rmdir()
            return True
        except (OSError, ValueError) as exc:
            log.error("Could not remove auto-capture transaction directory %s: %s", directory, exc)
            return False

    @classmethod
    def _validate_auto_capture_event_directory(
        cls,
        event_dir: Path,
        expected_manifest_rows: list[dict[str, str]],
        expected_classifier_snapshot: Optional[bytes],
    ) -> None:
        """Require the committed directory to match the WAL manifest exactly."""
        directory = Path(event_dir)
        if directory.is_symlink() or not directory.is_dir():
            raise ValueError("committed auto-capture event directory is missing")
        manifest_path = directory / "capture_manifest.csv"
        with open(manifest_path, newline="", encoding="utf-8") as stream:
            reader = csv.DictReader(stream)
            if tuple(reader.fieldnames or ()) != tuple(cls.AUTO_CAPTURE_MANIFEST_FIELDS):
                raise ValueError("auto-capture manifest schema does not match the WAL")
            actual_rows = [
                {
                    field: "" if row.get(field) is None else str(row.get(field))
                    for field in cls.AUTO_CAPTURE_MANIFEST_FIELDS
                }
                for row in reader
            ]
        if actual_rows != expected_manifest_rows:
            raise ValueError("auto-capture manifest rows do not match the WAL")
        for row in expected_manifest_rows:
            frame_name = row["frame_path"]
            if Path(frame_name).name != frame_name or not frame_name.lower().endswith(".bmp"):
                raise ValueError("auto-capture WAL contains an unsafe frame filename")
            frame_path = directory / frame_name
            if (
                frame_path.is_symlink()
                or not frame_path.is_file()
                or frame_path.stat().st_size <= 0
            ):
                raise ValueError("auto-capture WAL frame is missing or empty")
        snapshot_path = directory / "classifier_state.json"
        if expected_classifier_snapshot is None:
            if snapshot_path.exists() or snapshot_path.is_symlink():
                raise ValueError("auto-capture directory has an unjournaled classifier snapshot")
        elif (
            snapshot_path.is_symlink()
            or not snapshot_path.is_file()
            or snapshot_path.read_bytes() != expected_classifier_snapshot
        ):
            raise ValueError("auto-capture classifier snapshot does not match the WAL")

        expected_names = {
            "capture_manifest.csv",
            *(row["frame_path"] for row in expected_manifest_rows),
        }
        if expected_classifier_snapshot is not None:
            expected_names.add("classifier_state.json")
        actual_names = {child.name for child in directory.iterdir()}
        if actual_names != expected_names:
            raise ValueError("auto-capture event directory contains unjournaled files")

    @classmethod
    def _recover_auto_capture_transaction(cls, session_dir: Path) -> bool:
        """Finish or roll back one event-buffer/CSV transaction.

        The exact durable CSV row is the commit decision.  A matching row
        keeps (or promotes) its verified event directory; an absent row owns
        no data and therefore removes both staging and final directories.
        Malformed or conflicting state is retained fail-closed.
        """
        session_dir = Path(session_dir)
        marker = cls._auto_capture_transaction_path(session_dir)
        if not marker.exists():
            return True
        try:
            payload = json.loads(marker.read_text(encoding="utf-8"))
            if (
                not isinstance(payload, dict)
                or payload.get("schema_version")
                != cls.AUTO_CAPTURE_TRANSACTION_SCHEMA_VERSION
            ):
                raise ValueError("unsupported auto-capture WAL schema")
            row_payload = payload.get("row")
            if not isinstance(row_payload, dict) or set(row_payload) != set(cls.AUTO_CAPTURE_FIELDS):
                raise ValueError("auto-capture WAL row does not match the CSV schema")
            row = {
                field: "" if row_payload.get(field) is None else str(row_payload.get(field))
                for field in cls.AUTO_CAPTURE_FIELDS
            }
            event_idx = int(payload.get("event_idx"))
            if event_idx <= 0 or row["event_idx"] != str(event_idx):
                raise ValueError("auto-capture WAL event identity is inconsistent")
            buffer_count = int(row["buffer_count"])
            if buffer_count < 0:
                raise ValueError("auto-capture WAL buffer count is invalid")

            manifest_payload = payload.get("manifest_rows")
            if not isinstance(manifest_payload, list):
                raise ValueError("auto-capture WAL manifest is missing")
            manifest_rows: list[dict[str, str]] = []
            for manifest_row in manifest_payload:
                if (
                    not isinstance(manifest_row, dict)
                    or set(manifest_row) != set(cls.AUTO_CAPTURE_MANIFEST_FIELDS)
                ):
                    raise ValueError("auto-capture WAL manifest row is invalid")
                manifest_rows.append({
                    field: "" if manifest_row.get(field) is None else str(manifest_row.get(field))
                    for field in cls.AUTO_CAPTURE_MANIFEST_FIELDS
                })
            if len(manifest_rows) != buffer_count:
                raise ValueError("auto-capture WAL manifest count is inconsistent")
            frame_names = [manifest_row["frame_path"] for manifest_row in manifest_rows]
            if len(frame_names) != len(set(frame_names)):
                raise ValueError("auto-capture WAL contains duplicate frame filenames")

            snapshot_path_value = str(payload.get("classifier_snapshot_path", ""))
            snapshot_json_value = payload.get("classifier_snapshot_json", "")
            snapshot_hash = str(payload.get("classifier_snapshot_sha256", "")).lower()
            expected_classifier_snapshot: Optional[bytes] = None
            if snapshot_path_value:
                if snapshot_path_value != "classifier_state.json" or not buffer_count:
                    raise ValueError("auto-capture classifier snapshot path is invalid")
                if not isinstance(snapshot_json_value, str) or not snapshot_json_value:
                    raise ValueError("auto-capture classifier snapshot JSON is missing")
                expected_classifier_snapshot = snapshot_json_value.encode("utf-8")
                if (
                    len(snapshot_hash) != 64
                    or any(character not in "0123456789abcdef" for character in snapshot_hash)
                    or hashlib.sha256(expected_classifier_snapshot).hexdigest()
                    != snapshot_hash
                ):
                    raise ValueError("auto-capture classifier snapshot hash is invalid")
                snapshot_payload = json.loads(snapshot_json_value)
                if (
                    not isinstance(snapshot_payload, dict)
                    or int(snapshot_payload.get("event_idx", 0)) != event_idx
                ):
                    raise ValueError("auto-capture classifier snapshot identity is invalid")
            elif snapshot_json_value or snapshot_hash:
                raise ValueError("auto-capture classifier snapshot WAL fields conflict")

            staging_value = str(payload.get("staging_dir", ""))
            final_value = str(payload.get("final_dir", ""))
            staging_dir: Optional[Path] = None
            final_dir: Optional[Path] = None
            if buffer_count:
                staging_relative = Path(staging_value.replace("\\", "/"))
                final_relative = Path(final_value.replace("\\", "/"))
                expected_final = Path("frames") / f"auto_event_{event_idx:03d}"
                if (
                    final_relative != expected_final
                    or row["buffer_dir"] != expected_final.as_posix()
                    or staging_relative.parent != Path("frames")
                    or not staging_relative.name.startswith(
                        f".auto_event_{event_idx:03d}."
                    )
                    or not staging_relative.name.endswith(".pending")
                ):
                    raise ValueError("auto-capture WAL directory binding is invalid")
                staging_dir = cls._resolve_session_relative_path(
                    session_dir, staging_relative,
                )
                final_dir = cls._resolve_session_relative_path(
                    session_dir, final_relative,
                )
                if row["event_state"] != EVENT_STATE_PENDING or row["state_changed_at"]:
                    raise ValueError("buffered auto-capture WAL has invalid initial state")
            elif (
                staging_value
                or final_value
                or row["buffer_dir"]
                or manifest_rows
                or row["event_state"] != EVENT_STATE_AUTO_SKIPPED
                or not row["state_changed_at"]
            ):
                raise ValueError("empty auto-capture WAL has inconsistent paths or state")
            elif (session_dir / "frames" / f"auto_event_{event_idx:03d}").exists():
                raise ValueError(
                    "empty auto-capture WAL conflicts with an existing event directory"
                )

            csv_path = session_dir / "auto_capture_events.csv"
            rows: list[dict[str, str]] = []
            if csv_path.exists():
                with open(csv_path, newline="", encoding="utf-8") as stream:
                    reader = csv.DictReader(stream)
                    if tuple(reader.fieldnames or ()) != tuple(cls.AUTO_CAPTURE_FIELDS):
                        raise ValueError("auto-capture CSV schema does not match the WAL")
                    rows = [
                        {
                            field: "" if existing.get(field) is None else str(existing.get(field))
                            for field in cls.AUTO_CAPTURE_FIELDS
                        }
                        for existing in reader
                    ]
            same_index = [existing for existing in rows if existing["event_idx"] == str(event_idx)]
            if len(same_index) > 1:
                raise ValueError("auto-capture CSV contains a duplicate event index")
            committed = bool(same_index)
            if committed and same_index[0] != row:
                raise ValueError("auto-capture WAL conflicts with the committed CSV row")

            if committed and buffer_count:
                assert staging_dir is not None and final_dir is not None
                if final_dir.exists() and staging_dir.exists():
                    raise ValueError("auto-capture WAL has both staging and final directories")
                if not final_dir.exists():
                    if not staging_dir.exists():
                        raise ValueError("committed auto-capture row has no event directory")
                    cls._validate_auto_capture_event_directory(
                        staging_dir, manifest_rows, expected_classifier_snapshot,
                    )
                    os.replace(staging_dir, final_dir)
                cls._validate_auto_capture_event_directory(
                    final_dir, manifest_rows, expected_classifier_snapshot,
                )
            elif not committed:
                cleanup_ok = True
                for owned_dir in (staging_dir, final_dir):
                    if owned_dir is not None:
                        cleanup_ok = cls._remove_auto_capture_directory(owned_dir) and cleanup_ok
                if not cleanup_ok:
                    raise OSError("auto-capture rollback could not remove every directory")

            for temporary_path in csv_path.parent.glob(f".{csv_path.name}.*.tmp"):
                temporary_path.unlink(missing_ok=True)
            marker.unlink()
            return True
        except (OSError, TypeError, ValueError, json.JSONDecodeError) as exc:
            log.error(
                "Auto-capture transaction recovery failed for %s: %s",
                marker,
                exc,
            )
            return False

    def _recover_pending_live_label_transactions(self) -> bool:
        """Recover all image/provenance WALs in the configured log root."""
        try:
            live_markers = tuple(
                self._base_dir.glob(f"*/{self.LIVE_LABEL_TRANSACTION_FILE}")
            )
            calibration_markers = tuple(
                self._base_dir.glob(f"*/{self.CALIBRATION_TRANSACTION_FILE}")
            )
            auto_capture_markers = tuple(
                self._base_dir.glob(f"*/{self.AUTO_CAPTURE_TRANSACTION_FILE}")
            )
        except OSError as exc:
            log.error("Could not scan Equalizer transactions: %s", exc)
            return False
        recovered = True
        for marker in live_markers:
            if not self._recover_live_label_transaction(marker.parent):
                recovered = False
                log.error("Live-label transaction remains unresolved: %s", marker)
        for marker in calibration_markers:
            if not self._recover_calibration_transaction(marker.parent):
                recovered = False
                log.error("Calibration transaction remains unresolved: %s", marker)
        for marker in auto_capture_markers:
            if not self._recover_auto_capture_transaction(marker.parent):
                recovered = False
                log.error("Auto-capture transaction remains unresolved: %s", marker)
        return recovered

    def _close_live_label_stream(self) -> None:
        stream = self._live_label_file
        if stream is not None and not stream.closed:
            stream.close()
        self._live_label_file = None
        self._live_label_writer = None

    def _open_live_label_stream_for_append(self) -> bool:
        if self._session_dir is None:
            return False
        path = self._session_dir / "live_labels.csv"
        try:
            self._live_label_file = open(
                path, "a", newline="", encoding="utf-8",
            )
            self._live_label_writer = csv.DictWriter(
                self._live_label_file, fieldnames=self.LIVE_LABEL_FIELDS,
            )
            return True
        except OSError as exc:
            self._live_label_file = None
            self._live_label_writer = None
            log.error("Could not reopen Live-label CSV %s: %s", path, exc)
            return False

    def _close_auto_capture_stream(self) -> None:
        stream = self._auto_capture_file
        if stream is not None and not stream.closed:
            stream.close()
        self._auto_capture_file = None
        self._auto_capture_writer = None

    def _open_auto_capture_stream_for_append(self) -> bool:
        if self._session_dir is None:
            return False
        path = self._session_dir / "auto_capture_events.csv"
        try:
            self._auto_capture_file = open(
                path, "a", newline="", encoding="utf-8",
            )
            self._auto_capture_writer = csv.DictWriter(
                self._auto_capture_file, fieldnames=self.AUTO_CAPTURE_FIELDS,
            )
            return True
        except OSError as exc:
            self._auto_capture_file = None
            self._auto_capture_writer = None
            log.error("Could not reopen auto-capture CSV %s: %s", path, exc)
            return False

    def _auto_capture_stream_ready(self) -> bool:
        stream = self._auto_capture_file
        return bool(
            self._auto_capture_writer is not None
            and stream is not None
            and not stream.closed
        )

    @staticmethod
    def _remove_live_label_transaction_files(*paths: Path) -> bool:
        """Roll back artifacts, clearing the WAL only after cleanup succeeds."""
        if not paths:
            return True
        *artifacts, marker = (Path(path) for path in paths)
        removed = True
        for path in artifacts:
            try:
                path.unlink(missing_ok=True)
            except OSError as exc:
                removed = False
                log.error("Could not clean Live-label transaction path %s: %s", path, exc)
        if not removed:
            log.error("Retaining Live-label WAL after incomplete rollback: %s", marker)
            return False
        try:
            marker.unlink(missing_ok=True)
        except OSError as exc:
            log.error("Could not clear rolled-back Live-label WAL %s: %s", marker, exc)
            return False
        return removed

    def __init__(self, base_dir: str = "logs/growths"):
        self._base_dir = Path(base_dir)
        self._filename_prefix: str = "growth"
        self._session_dir: Optional[Path] = None
        self._sensor_file = None
        self._sensor_writer = None
        self._commit_file = None
        self._commit_writer = None
        self._auto_capture_file = None
        self._auto_capture_writer = None
        self._heartbeat_file = None
        self._heartbeat_writer = None
        self._set_change_file = None
        self._set_change_writer = None
        self._manual_event_file = None
        self._manual_event_writer = None
        self._rheed_view_event_file = None
        self._rheed_view_event_writer = None
        self._live_label_file = None
        self._live_label_writer = None
        self._equalizer_calibration_file = None
        self._calibration_journal_trusted = True
        self._commit_counter = 0
        self._heartbeat_counter = 0
        self._set_change_counter = 0
        self._manual_event_counter = 0
        self._rheed_view_event_counter = 0
        self._live_label_counter = 0
        self._equalizer_calibration_event_counter = 0
        self._basis_bundle_ids_recorded: set[str] = set()
        # A fresh runtime identifier is intentionally not persisted across
        # process restarts.  Blind-label provenance is valid only after this
        # runtime explicitly attaches to the session and enters blind mode.
        self._human_labeling_runtime_id = uuid.uuid4().hex
        self._human_labeling_state_trusted = False
        self._sensor_row_counter = 0
        self._session_start: Optional[datetime] = None
        self._entries: list[dict] = []  # Accumulated entries for export
        self._recover_pending_live_label_transactions()

    @property
    def active(self) -> bool:
        """True while CSV files are open and accepting writes."""
        return self._sensor_writer is not None

    @property
    def session_dir(self) -> Optional[Path]:
        return self._session_dir

    def set_base_dir(self, base_dir: str | Path) -> bool:
        """Apply a runtime log root and recover its interrupted transactions.

        Growth Monitor reads the production path from the GUI after this
        logger is constructed.  Scanning only the constructor default would
        miss WALs under that configured drive on every restart.
        """
        if self.active:
            raise RuntimeError("Cannot change the log root during an active session")
        self._base_dir = Path(base_dir)
        return self._recover_pending_live_label_transactions()

    def start_session(self, sample_id: str):
        """Create session directory and open CSV files."""
        # Clear any previous session data
        self._session_dir = None
        self._entries = []
        tag = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_id = sample_id.strip().replace(" ", "_") or "unnamed"
        prefix = self._filename_prefix or "growth"
        self._session_dir = self._base_dir / f"{prefix}_{safe_id}_{tag}"
        self._session_dir.mkdir(parents=True, exist_ok=True)
        (self._session_dir / "frames").mkdir(exist_ok=True)
        self._calibration_journal_trusted = True
        self._basis_bundle_ids_recorded.clear()
        if not self._initialize_human_labeling_state():
            log.error(
                "Human blind-labeling proof state could not be initialized; "
                "this session will produce assisted labels only"
            )

        sensor_path = self._session_dir / "sensor_log.csv"
        self._sensor_file = open(sensor_path, "w", newline="")
        self._sensor_writer = csv.DictWriter(
            self._sensor_file, fieldnames=self.SENSOR_FIELDS,
        )
        self._sensor_writer.writeheader()

        commit_path = self._session_dir / "commit_log.csv"
        self._commit_file = open(commit_path, "w", newline="")
        self._commit_writer = csv.DictWriter(
            self._commit_file, fieldnames=self.COMMIT_FIELDS,
        )
        self._commit_writer.writeheader()

        auto_capture_path = self._session_dir / "auto_capture_events.csv"
        self._auto_capture_file = open(auto_capture_path, "w", newline="")
        self._auto_capture_writer = csv.DictWriter(
            self._auto_capture_file, fieldnames=self.AUTO_CAPTURE_FIELDS,
        )
        self._auto_capture_writer.writeheader()

        heartbeat_path = self._session_dir / "heartbeat_log.csv"
        self._heartbeat_file = open(heartbeat_path, "w", newline="")
        self._heartbeat_writer = csv.DictWriter(
            self._heartbeat_file, fieldnames=self.HEARTBEAT_FIELDS,
        )
        self._heartbeat_writer.writeheader()

        set_change_path = self._session_dir / "set_change_events.csv"
        self._set_change_file = open(set_change_path, "w", newline="")
        self._set_change_writer = csv.DictWriter(
            self._set_change_file, fieldnames=self.SET_CHANGE_FIELDS,
        )
        self._set_change_writer.writeheader()

        manual_event_path = self._session_dir / "manual_events.csv"
        self._manual_event_file = open(manual_event_path, "w", newline="")
        self._manual_event_writer = csv.DictWriter(
            self._manual_event_file, fieldnames=self.MANUAL_EVENT_FIELDS,
        )
        self._manual_event_writer.writeheader()
        # Flush the header immediately — downstream scrubber-timeline
        # tooling may open the file for read while the session is running,
        # and DictWriter buffers by default. Cost is negligible (one
        # header line, one flush, per session).
        self._manual_event_file.flush()

        rheed_view_event_path = self._session_dir / "rheed_view_events.csv"
        self._rheed_view_event_file = open(
            rheed_view_event_path, "w", newline="",
        )
        self._rheed_view_event_writer = csv.DictWriter(
            self._rheed_view_event_file,
            fieldnames=self.RHEED_VIEW_EVENT_FIELDS,
        )
        self._rheed_view_event_writer.writeheader()
        # Flush immediately so acquisition-state consumers can attach while
        # the session is running, even before the first transition occurs.
        self._rheed_view_event_file.flush()

        live_label_path = self._session_dir / "live_labels.csv"
        self._live_label_file = open(
            live_label_path, "w", newline="", encoding="utf-8",
        )
        self._live_label_writer = csv.DictWriter(
            self._live_label_file, fieldnames=self.LIVE_LABEL_FIELDS,
        )
        self._live_label_writer.writeheader()
        self._live_label_file.flush()

        # Append-only journal.  Every accepted calibration and subsequent
        # invalidation is retained as a separate JSON line so a label can be
        # reconstructed without consulting mutable GUI state.
        calibration_path = self._session_dir / "equalizer_calibrations.jsonl"
        self._equalizer_calibration_file = open(
            calibration_path, "a", encoding="utf-8", newline="\n",
        )

        self._commit_counter = 0
        self._heartbeat_counter = 0
        self._set_change_counter = 0
        self._manual_event_counter = 0
        self._rheed_view_event_counter = 0
        self._live_label_counter = 0
        self._equalizer_calibration_event_counter = 0
        self._sensor_row_counter = 0
        self._session_start = datetime.now()
        self._entries = []

    def log_sensors(
        self, pyro_temp, elapsed_s,
        v_set=None, v_actual=None, i_set=None, i_actual=None,
        chamber_pressure_mbar=None,
        pyro_temp_std=None, pyro_temp_n=None,
        # Elog-direct extensions (Jun 23 2026 — EvapControl mode="elog").
        # All optional; blank in the CSV when None. Order matches
        # EvapControlState field order so the call site can pass through
        # state attributes one-for-one.
        substrate_temp_pv_C=None, substrate_temp_setpoint_C=None,
        cell_HTEC2_pv_C=None,
        cell_Y_pv_C=None, cell_Sr_pv_C=None,
        cell_Eu_pv_C=None, cell_Er_pv_C=None,
        plasma_dc_bias_V=None, plasma_forward_W=None,
        plasma_reflected_W=None,
        # ADS-mode extensions (Jul 23 2026 — MistralWorker mode="ads").
        # Full read() dict from MistralAdsClient; None in O-MBE sessions.
        ads_cells=None,
    ):
        """Append a row to sensor_log.csv. All values may be None.

        ``pyro_temp_std`` and ``pyro_temp_n`` capture the per-poll
        statistical spread when the pyrometer worker takes multiple
        sub-readings per cycle. Empty strings if not provided.

        The ``substrate_*``, ``cell_*``, and ``plasma_*`` kwargs are
        populated only when EvapControl is in ``elog`` mode (reading the
        .elo binary log directly). In ``screengrab`` mode they default
        to None and the columns are blank.
        """
        if not self._sensor_writer:
            return

        def _f(val, places):
            return f"{val:.{places}f}" if val is not None else ""

        def _sci(val):
            return f"{val:.3e}" if val is not None else ""

        self._sensor_writer.writerow({
            "timestamp": datetime.now().isoformat(),
            "elapsed_s": f"{elapsed_s:.2f}",
            "pyrometer_temp_C":     _f(pyro_temp, 1),
            "pyrometer_temp_std_C": _f(pyro_temp_std, 2),
            "pyrometer_temp_n":     pyro_temp_n if pyro_temp_n is not None else "",
            "mistral_v_set_V":    _f(v_set, 3),
            "mistral_v_actual_V": _f(v_actual, 3),
            "mistral_i_set_A":    _f(i_set, 3),
            "mistral_i_actual_A": _f(i_actual, 3),
            "chamber_pressure_mbar": _sci(chamber_pressure_mbar),
            # Elog-direct fields. Temps to 1 dp (TemperaSure-style),
            # plasma DC bias to 1 dp, plasma powers to 1 dp.
            "substrate_temp_pv_C":       _f(substrate_temp_pv_C, 1),
            "substrate_temp_setpoint_C": _f(substrate_temp_setpoint_C, 1),
            "cell_HTEC2_pv_C": _f(cell_HTEC2_pv_C, 1),
            "cell_Y_pv_C":     _f(cell_Y_pv_C, 1),
            "cell_Sr_pv_C":    _f(cell_Sr_pv_C, 1),
            "cell_Eu_pv_C":    _f(cell_Eu_pv_C, 1),
            "cell_Er_pv_C":    _f(cell_Er_pv_C, 1),
            "plasma_dc_bias_V":   _f(plasma_dc_bias_V, 1),
            "plasma_forward_W":   _f(plasma_forward_W, 1),
            "plasma_reflected_W": _f(plasma_reflected_W, 1),
            # ADS cell temps — populated from ads_cells dict when mode="ads"
            **{
                f"cell{i}_T_C": _f(
                    ads_cells.get(f"cell{i}_T") if ads_cells else None, 1
                )
                for i in range(1, 8)
            },
        })
        self._sensor_file.flush()
        self._sensor_row_counter += 1

    def log_commit(self, entry: dict):
        """Append a row to commit_log.csv and accumulate for export."""
        if not self._commit_writer:
            return
        row = {field: entry.get(field, "") for field in self.COMMIT_FIELDS}
        self._commit_writer.writerow(row)
        self._commit_file.flush()
        self._entries.append(entry)

    def save_heartbeat_frame(
        self, frame: np.ndarray, timestamp: str = ""
    ) -> str:
        """Save a heartbeat anchor frame as ``heartbeat_NNN_HHMMSS.bmp``.

        BMP matches Justin's training-data format (Jul 9 2026 switch;
        was .png before). Same quality gate as ``save_frame``; failed
        frames return "" and are not counted toward the heartbeat counter.
        """
        if self._session_dir is None:
            return ""

        try:
            from drivers.frame_quality import check_frame_quality
            qa = check_frame_quality(frame)
            if not qa.passed:
                import sys
                print(
                    f"[GrowthLogger] heartbeat frame rejected: {qa.reason}",
                    file=sys.stderr,
                    flush=True,
                )
                return ""
        except ImportError:
            pass

        self._heartbeat_counter += 1
        ts = timestamp or datetime.now().strftime("%H%M%S")
        # .bmp per Jul 9 2026 switch — matches Justin's training data
        # format. PIL / cv2 infer format from the extension; no other
        # code change needed.
        fname = f"heartbeat_{self._heartbeat_counter:03d}_{ts}.bmp"
        path = self._session_dir / "frames" / fname

        try:
            from PIL import Image
            Image.fromarray(frame).save(str(path))
        except ImportError:
            try:
                import cv2
                cv2.imwrite(str(path), cv2.cvtColor(frame, cv2.COLOR_RGB2BGR))
            except ImportError:
                return ""

        return str(path)

    def log_set_change_event(
        self,
        elapsed_s: float,
        channel: str,
        old_value: float,
        new_value: float,
        pyro_temp: Optional[float] = None,
    ):
        """Append a row to set_change_events.csv.

        Pairs operator setpoint changes (Set Voltage / Set Current button
        presses on the MISTRAL GUI) with timestamps, so the model can
        learn from when V/I are deliberately adjusted vs drifting.
        """
        if not self._set_change_writer:
            return
        self._set_change_counter += 1
        self._set_change_writer.writerow({
            "timestamp": datetime.now().isoformat(),
            "elapsed_s": f"{elapsed_s:.2f}",
            "event_idx": self._set_change_counter,
            "channel": channel,
            "old_value": f"{old_value:.4f}",
            "new_value": f"{new_value:.4f}",
            "delta": f"{new_value - old_value:+.4f}",
            "pyrometer_temp_C": (
                f"{pyro_temp:.1f}" if pyro_temp is not None else ""
            ),
        })
        self._set_change_file.flush()

    def record_manual_event(
        self,
        elapsed_s: float,
        pyro_temp: Optional[float] = None,
        voltage_V: Optional[float] = None,
        current_A: Optional[float] = None,
        psu_source: str = "none",
        frame: Optional[np.ndarray] = None,
        note: str = "",
        capture_metadata: Optional[dict] = None,
    ) -> int:
        """Append a grower-marked event to manual_events.csv.

        Sub-second reconstruction transitions surfaced at the Jul 10 2026
        group meeting as a hard UX requirement — labeling takes ~30 s but
        reconstructions can flip in ~1 s. Manual events decouple "grower
        saw something NOW" from "grower labels retrospectively later".
        Every click writes one row; the scrubber (workstream #3) reads
        this file to render manual markers on the timeline.

        When ``frame`` is provided, it's saved as
        ``manual_event_NNN_HHMMSS.bmp`` alongside ``entry_*`` and
        ``heartbeat_*`` under ``frames/`` — the scrubber can attach a
        per-mark preview. Frame save is best-effort: the CSV row is
        written even if PIL/cv2 aren't importable (grower's click is
        stronger evidence of intent than a working image encoder).

        Returns the 1-indexed event_idx so callers can update UI
        counters without re-reading the file. Returns 0 (silent no-op)
        if no session is active — mirrors log_commit's contract.
        """
        if not self._manual_event_writer:
            return 0

        self._manual_event_counter += 1
        idx = self._manual_event_counter

        # Save the frame first — a failed frame save must not block the
        # row from landing (grower's click is evidence-of-intent). Same
        # BMP + name-taxonomy pattern as save_frame + save_heartbeat_frame.
        frame_path = ""
        if frame is not None and self._session_dir is not None:
            ts = datetime.now().strftime("%H%M%S")
            fname = f"manual_event_{idx:03d}_{ts}.bmp"
            path = self._session_dir / "frames" / fname
            try:
                from PIL import Image
                Image.fromarray(frame).save(str(path))
                frame_path = str(path)
            except ImportError:
                try:
                    import cv2
                    cv2.imwrite(
                        str(path), cv2.cvtColor(frame, cv2.COLOR_RGB2BGR),
                    )
                    frame_path = str(path)
                except ImportError:
                    pass

        self._manual_event_writer.writerow({
            "timestamp": datetime.now().isoformat(),
            "elapsed_s": f"{elapsed_s:.2f}",
            "event_idx": idx,
            "pyrometer_temp_C": (
                f"{pyro_temp:.1f}" if pyro_temp is not None else ""
            ),
            "voltage_V": (
                f"{voltage_V:.3f}" if voltage_V is not None else ""
            ),
            "current_A": (
                f"{current_A:.3f}" if current_A is not None else ""
            ),
            "psu_source": psu_source,
            "frame_path": frame_path,
            "note": note,
            **self._capture_columns(capture_metadata),
        })
        self._manual_event_file.flush()
        return idx

    def record_rheed_view_event(
        self,
        event_type: str,
        elapsed_s: float,
        *,
        state_snapshot: Optional[Mapping[str, object]] = None,
        realignment_id: Optional[int] = None,
        previous_view_segment_id: Optional[int] = None,
        frame_role: str = "",
        frame: Optional[np.ndarray] = None,
        note: str = "",
        capture_metadata: Optional[dict] = None,
        labeler: str = "",
        confidence: str = "",
    ) -> int:
        """Append a structured acquisition/QC transition.

        ``state_snapshot`` is the post-event acquisition state. Recognized
        keys are ``view_segment_id``, ``visual_history_generation``,
        ``gun_aligned``,
        ``history_frame_count``, ``history_required``, ``history_ready``,
        ``qc_reject``, ``qc_reason``, ``labeler``, ``confidence``, and
        ``prediction_actionable``.
        ``realignment_id`` and ``previous_view_segment_id`` may be passed
        explicitly or supplied in the snapshot.

        The optional frame is evidence captured at the transition boundary,
        not an image to be discarded. ``frame_role`` should therefore say
        what it represents (for example ``"pre_realign"`` or
        ``"post_realign"``). Saving is best-effort, while the structured
        event row is always written.

        Returns the 1-indexed event id, or 0 when no session is active.
        Unknown event types raise ``ValueError`` before any counter or file
        mutation.
        """
        if event_type not in RHEED_VIEW_EVENT_TYPES:
            allowed = ", ".join(sorted(RHEED_VIEW_EVENT_TYPES))
            raise ValueError(
                f"Unsupported RHEED view event_type {event_type!r}; "
                f"expected one of: {allowed}"
            )
        if state_snapshot is None:
            snapshot: Mapping[str, object] = {}
        elif isinstance(state_snapshot, Mapping):
            snapshot = state_snapshot
        else:
            raise TypeError("state_snapshot must be a mapping or None")

        if not self._rheed_view_event_writer:
            return 0

        self._rheed_view_event_counter += 1
        idx = self._rheed_view_event_counter

        frame_path = ""
        if frame is not None and self._session_dir is not None:
            ts = datetime.now().strftime("%H%M%S")
            fname = (
                f"rheed_view_event_{idx:03d}_{event_type}_{ts}.bmp"
            )
            path = self._session_dir / "frames" / fname
            try:
                from PIL import Image
                Image.fromarray(frame).save(str(path))
                frame_path = str(path)
            except ImportError:
                try:
                    import cv2
                    cv2.imwrite(
                        str(path), cv2.cvtColor(frame, cv2.COLOR_RGB2BGR),
                    )
                    frame_path = str(path)
                except ImportError:
                    pass

        def _value(key: str) -> object:
            value = snapshot.get(key, "")
            if value is None:
                return ""
            if isinstance(value, bool):
                return "True" if value else "False"
            return value

        realignment_value: object = realignment_id
        if realignment_value is None:
            realignment_value = _value("realignment_id")
        previous_segment_value: object = previous_view_segment_id
        if previous_segment_value is None:
            previous_segment_value = _value("previous_view_segment_id")

        self._rheed_view_event_writer.writerow({
            "timestamp": datetime.now().isoformat(),
            "elapsed_s": f"{elapsed_s:.2f}",
            "event_idx": idx,
            "event_type": event_type,
            "realignment_id": (
                "" if realignment_value is None else realignment_value
            ),
            "previous_view_segment_id": (
                "" if previous_segment_value is None
                else previous_segment_value
            ),
            "view_segment_id": _value("view_segment_id"),
            "visual_history_generation": _value(
                "visual_history_generation"
            ),
            "gun_aligned": _value("gun_aligned"),
            "history_frame_count": _value("history_frame_count"),
            "history_required": _value("history_required"),
            "history_ready": _value("history_ready"),
            "qc_reject": _value("qc_reject"),
            "qc_reason": _value("qc_reason"),
            "labeler": labeler or _value("labeler"),
            "confidence": confidence or _value("confidence"),
            "prediction_actionable": _value("prediction_actionable"),
            "frame_role": frame_role,
            "frame_path": frame_path,
            "note": note,
            **self._capture_columns(capture_metadata),
        })
        self._rheed_view_event_file.flush()
        return idx

    def _calibration_journal_path(self) -> Optional[Path]:
        if self._session_dir is None:
            return None
        return self._session_dir / "equalizer_calibrations.jsonl"

    def _calibration_failure_marker_path(self) -> Optional[Path]:
        path = self._calibration_journal_path()
        return None if path is None else path.with_suffix(".untrusted")

    def _write_calibration_failure_marker(self, reason: str) -> bool:
        """Durably create the marker that makes journal replay fail closed."""
        marker = self._calibration_failure_marker_path()
        if marker is None:
            return False
        try:
            with open(marker, "w", encoding="utf-8", newline="\n") as stream:
                stream.write(
                    datetime.now(timezone.utc).isoformat()
                    + " " + str(reason).strip() + "\n"
                )
                stream.flush()
                os.fsync(stream.fileno())
            return True
        except OSError as exc:
            log.critical(
                "Could not persist Equalizer journal failure marker %s: %s",
                marker,
                exc,
            )
            return False

    def _mark_calibration_journal_untrusted(self, reason: str) -> None:
        self._calibration_journal_trusted = False
        self._write_calibration_failure_marker(reason)
        log.error("Equalizer calibration journal is untrusted: %s", reason)

    def _append_calibration_event(self, event: dict) -> bool:
        """Durably append during or after a session, failing future reads closed.

        A durable marker is written *before* the JSONL append.  It is removed
        only after the journal line has been flushed and fsynced.  A crash at
        any intermediate point therefore makes replay reject the journal.
        """
        path = self._calibration_journal_path()
        marker = self._calibration_failure_marker_path()
        if (
            path is None
            or not self._calibration_journal_trusted
            or (marker is not None and marker.exists())
        ):
            return False

        # Refuse to append to a journal that is already malformed.  Replay
        # persists its own untrusted marker when validation fails.
        self._replay_calibration_journal()
        if not self._calibration_journal_trusted:
            return False
        try:
            line = json.dumps(
                event, sort_keys=True, separators=(",", ":"), allow_nan=False,
            ) + "\n"
        except (TypeError, ValueError) as exc:
            self._mark_calibration_journal_untrusted(
                f"calibration event serialization failed: {exc}"
            )
            return False
        if marker is None or not self._write_calibration_failure_marker(
            "append in progress"
        ):
            self._calibration_journal_trusted = False
            return False

        owned_stream = None
        stream = self._equalizer_calibration_file
        append_error: Optional[BaseException] = None
        try:
            if stream is None or stream.closed:
                owned_stream = open(path, "a", encoding="utf-8", newline="\n")
                stream = owned_stream
            stream.write(line)
            stream.flush()
            os.fsync(stream.fileno())
        except (OSError, ValueError, TypeError) as exc:
            append_error = exc
        finally:
            if owned_stream is not None:
                try:
                    owned_stream.close()
                except OSError as exc:
                    append_error = append_error or exc
        if append_error is not None:
            self._mark_calibration_journal_untrusted(
                f"calibration journal append failed: {append_error}"
            )
            return False
        try:
            marker.unlink()
        except OSError as exc:
            self._mark_calibration_journal_untrusted(
                f"calibration journal commit marker could not be cleared: {exc}"
            )
            return False
        self._calibration_journal_trusted = True
        self._equalizer_calibration_event_counter += 1
        return True

    def record_calibration(
        self,
        calibration: CalibrationRecord,
        evidence_snapshot: Optional[RheedFrameSnapshot] = None,
        basis_bundle: Optional[BasisBundle] = None,
    ) -> bool:
        """Persist evidence, then append a complete accepted calibration."""
        if self._session_dir is None:
            return False
        if not isinstance(calibration, CalibrationRecord):
            raise TypeError("calibration must be a CalibrationRecord")
        if not calibration.grower_accepted or calibration.invalidated_reason:
            raise ValueError("Only an accepted, active calibration may be recorded")
        if not isinstance(basis_bundle, BasisBundle):
            raise TypeError("Accepted calibration requires its BasisBundle")
        if basis_bundle.bundle_id != calibration.basis_bundle_id:
            raise ValueError("Calibration basis bundle does not match its manifest")
        basis_manifest = BasisBundle.validate_manifest_dict(
            basis_bundle.to_manifest_dict()
        )
        if (
            not calibration.gun_aligned
            or calibration.realignment_active
            or not calibration.session_id
            or calibration.view_segment_id is None
        ):
            raise ValueError("Calibration is not bound to a stable session view")
        self._validate_capture_identity(
            captured_at_utc=calibration.captured_at_utc,
            capture_sequence=calibration.capture_sequence,
            received_monotonic_ns=calibration.received_monotonic_ns,
            capture_backend=calibration.capture_backend,
            source_hwnd=calibration.source_hwnd,
            capture_geometry_id=calibration.capture_geometry_id,
            camera_width=calibration.camera_width,
            camera_height=calibration.camera_height,
        )
        if (
            self._session_dir is None
            or calibration.session_id != self._session_dir.name
        ):
            raise ValueError("Calibration belongs to a different session")
        if evidence_snapshot is None:
            raise ValueError("Accepted calibration requires its frozen evidence frame")
        self._validate_equalizer_context(calibration, evidence_snapshot)
        exact_identity = (
            ("captured_at_utc", calibration.captured_at_utc,
             evidence_snapshot.captured_at_utc),
            ("capture_sequence", calibration.capture_sequence,
             evidence_snapshot.capture_sequence),
            ("received_monotonic_ns", calibration.received_monotonic_ns,
             evidence_snapshot.received_monotonic_ns),
        )
        for label, expected, actual in exact_identity:
            if expected != actual:
                raise ValueError(
                    f"Orientation evidence {label} does not match calibration"
                )
        if self.get_historical_calibration(calibration.calibration_id) is not None:
            return False

        try:
            evidence_png = evidence_snapshot.orientation_evidence_png()
        except Exception as exc:  # PIL encoder boundary
            log.error("Failed to encode calibration orientation evidence: %s", exc)
            return False
        evidence_hash = hashlib.sha256(evidence_png).hexdigest()
        if evidence_hash != calibration.orientation_evidence_sha256.lower():
            log.error(
                "Calibration orientation evidence hash mismatch for %s",
                calibration.calibration_id,
            )
            return False
        relative_evidence_path = Path(
            calibration.orientation_evidence_path.replace("\\", "/")
        )
        if relative_evidence_path.is_absolute() or ".." in relative_evidence_path.parts:
            raise ValueError("Calibration orientation evidence path is unsafe")
        evidence_path = self._session_dir / relative_evidence_path
        event = {
            "journal_schema_version": 1,
            "event": "accepted",
            "recorded_at_utc": datetime.now(timezone.utc).isoformat(),
            "calibration": calibration.to_json_dict(),
            "orientation_evidence_path": calibration.orientation_evidence_path,
        }
        if calibration.basis_bundle_id not in self._basis_bundle_ids_recorded:
            event["basis_bundle_manifest"] = basis_manifest

        transaction_path = self._calibration_transaction_path(self._session_dir)
        if transaction_path.exists():
            log.error(
                "Calibration acceptance blocked by unresolved transaction %s",
                transaction_path,
            )
            return False
        if evidence_path.exists():
            log.error(
                "Unjournaled orientation evidence already exists for %s: %s",
                calibration.calibration_id,
                evidence_path,
            )
            return False

        wal_payload = {
            "schema_version": self.CALIBRATION_TRANSACTION_SCHEMA_VERSION,
            "calibration_id": calibration.calibration_id,
            "evidence_path": calibration.orientation_evidence_path,
            "evidence_sha256": evidence_hash,
            "event": event,
        }
        try:
            wal_bytes = json.dumps(
                wal_payload,
                sort_keys=True,
                separators=(",", ":"),
                allow_nan=False,
            ).encode("utf-8")
        except (TypeError, ValueError) as exc:
            log.error("Could not serialize calibration transaction: %s", exc)
            return False
        if not self._atomic_write_bytes(transaction_path, wal_bytes):
            return False
        if not self._atomic_write_bytes(evidence_path, evidence_png):
            try:
                transaction_path.unlink(missing_ok=True)
            except OSError as exc:
                log.error(
                    "Could not clear failed calibration WAL %s: %s",
                    transaction_path,
                    exc,
                )
            return False

        appended = self._append_calibration_event(event)
        if appended:
            self._basis_bundle_ids_recorded.add(calibration.basis_bundle_id)
            try:
                transaction_path.unlink()
            except OSError as exc:
                # The journal and evidence are already durable.  Startup
                # recovery verifies both before clearing this WAL.
                log.error(
                    "Committed calibration but could not clear WAL %s: %s",
                    transaction_path,
                    exc,
                )
        else:
            try:
                evidence_path.unlink(missing_ok=True)
            except OSError as exc:
                log.error("Could not remove orphan orientation evidence %s: %s", evidence_path, exc)
                return False
            try:
                transaction_path.unlink(missing_ok=True)
            except OSError as exc:
                log.error("Could not clear rolled-back calibration WAL %s: %s", transaction_path, exc)
        return appended

    def record_calibration_invalidation(
        self,
        calibration: CalibrationRecord,
        reason: str,
    ) -> bool:
        """Append an invalidation without altering the accepted journal line."""
        if self._session_dir is None:
            return False
        if not isinstance(calibration, CalibrationRecord):
            raise TypeError("calibration must be a CalibrationRecord")
        reason = str(reason).strip()
        if not reason:
            raise ValueError("Calibration invalidation requires a reason")
        journaled = self.get_historical_calibration(
            calibration.calibration_id,
        )
        if journaled is None:
            return False
        if journaled.to_json_dict() != calibration.to_json_dict():
            raise ValueError("Calibration invalidation record does not match journal")
        # Repeated invalidation requests are normal during overlapping GUI
        # lifecycle signals.  Preserve exact historical identity validation
        # above, but do not manufacture duplicate journal events once inactive.
        if self.get_calibration(calibration.calibration_id) is None:
            return True
        invalidated = calibration.invalidated(reason)
        event = {
            "journal_schema_version": 1,
            "event": "invalidated",
            "recorded_at_utc": datetime.now(timezone.utc).isoformat(),
            "calibration": invalidated.to_json_dict(),
        }
        return self._append_calibration_event(event)

    def _replay_calibration_journal(
        self,
    ) -> tuple[dict[str, CalibrationRecord], dict[str, CalibrationRecord]]:
        """Return active and historically accepted records, or two empties."""
        path = self._calibration_journal_path()
        marker = self._calibration_failure_marker_path()
        if (
            path is None
            or not path.exists()
            or not self._calibration_journal_trusted
            or (marker is not None and marker.exists())
        ):
            return {}, {}
        active: dict[str, CalibrationRecord] = {}
        accepted_history: dict[str, CalibrationRecord] = {}
        basis_manifests: dict[str, dict] = {}

        def _fail(reason: str) -> tuple[dict, dict]:
            self._basis_bundle_ids_recorded.clear()
            self._mark_calibration_journal_untrusted(reason)
            return {}, {}

        try:
            with open(path, "r", encoding="utf-8") as stream:
                for line_number, raw_line in enumerate(stream, start=1):
                    if not raw_line.strip():
                        continue
                    event = json.loads(raw_line)
                    if not isinstance(event, dict):
                        return _fail(f"journal line {line_number} is not an object")
                    if type(event.get("journal_schema_version")) is not int or (
                        event.get("journal_schema_version") != 1
                    ):
                        return _fail(
                            f"journal line {line_number} has unsupported schema"
                        )
                    normalize_utc_timestamp(
                        event.get("recorded_at_utc", ""),
                        field_name=f"journal line {line_number} recorded_at_utc",
                    )
                    record = CalibrationRecord.from_json_dict(
                        event["calibration"],
                    )
                    event_type = event.get("event")
                    if event_type == "accepted":
                        if record.grower_accepted and not record.invalidated_reason:
                            manifest_value = event.get("basis_bundle_manifest")
                            if manifest_value is not None:
                                manifest = BasisBundle.validate_manifest_dict(
                                    manifest_value,
                                )
                                manifest_id = str(manifest["bundle_id"])
                                if manifest_id in basis_manifests:
                                    return _fail(
                                        "duplicate basis bundle manifest at line "
                                        f"{line_number}"
                                    )
                                if manifest_id != record.basis_bundle_id:
                                    return _fail(
                                        "basis bundle manifest does not match calibration "
                                        f"at line {line_number}"
                                    )
                                basis_manifests[manifest_id] = manifest
                            if record.basis_bundle_id not in basis_manifests:
                                return _fail(
                                    "calibration references an unrecorded basis bundle "
                                    f"at line {line_number}"
                                )
                            if record.calibration_id in accepted_history:
                                return _fail(
                                    f"duplicate accepted calibration at line {line_number}"
                                )
                            evidence_path = (
                                self._session_dir
                                / Path(record.orientation_evidence_path)
                            )
                            try:
                                evidence_digest = hashlib.sha256(
                                    evidence_path.read_bytes()
                                ).hexdigest()
                            except OSError as exc:
                                return _fail(
                                    "orientation evidence is unreadable at line "
                                    f"{line_number}: {exc}"
                                )
                            if (
                                evidence_digest
                                != record.orientation_evidence_sha256.lower()
                            ):
                                return _fail(
                                    "orientation evidence hash mismatch at line "
                                    f"{line_number}"
                                )
                            active[record.calibration_id] = record
                            accepted_history[record.calibration_id] = record
                        else:
                            return _fail(
                                f"invalid accepted calibration at line {line_number}"
                            )
                    elif event_type == "invalidated":
                        accepted = accepted_history.get(record.calibration_id)
                        if (
                            accepted is None
                            or record.grower_accepted
                            or not record.invalidated_reason
                            or accepted.invalidated(record.invalidated_reason).to_json_dict()
                            != record.to_json_dict()
                        ):
                            return _fail(
                                f"invalid calibration invalidation at line {line_number}"
                            )
                        active.pop(record.calibration_id, None)
                    else:
                        return _fail(
                            f"unknown calibration event at line {line_number}"
                        )
        except (OSError, KeyError, TypeError, ValueError, json.JSONDecodeError) as exc:
            return _fail(f"calibration journal replay failed: {exc}")
        self._basis_bundle_ids_recorded = set(basis_manifests)
        return active, accepted_history

    def read_calibrations(self) -> dict[str, CalibrationRecord]:
        """Replay the JSONL journal and return accepted, non-invalid records."""
        active, _ = self._replay_calibration_journal()
        return active

    def read_historical_calibrations(self) -> dict[str, CalibrationRecord]:
        """Return accepted records usable only for exact historical frames."""
        _, history = self._replay_calibration_journal()
        return history

    def get_calibration(
        self,
        calibration_id: str,
    ) -> Optional[CalibrationRecord]:
        """Return one currently valid journaled calibration by exact ID."""
        return self.read_calibrations().get(str(calibration_id))

    def get_historical_calibration(
        self,
        calibration_id: str,
    ) -> Optional[CalibrationRecord]:
        return self.read_historical_calibrations().get(str(calibration_id))

    def find_compatible_calibration(
        self,
        *,
        session_id: str,
        view_segment_id: int,
        visual_history_generation: int,
        basis_bundle_id: str,
        capture_backend: Optional[str] = None,
        source_hwnd: Optional[int] = None,
        capture_geometry_id: Optional[str] = None,
        camera_width: Optional[int] = None,
        camera_height: Optional[int] = None,
    ) -> Optional[CalibrationRecord]:
        """Return the newest active record matching exact lifecycle context."""
        matches: list[CalibrationRecord] = []
        for record in self.read_calibrations().values():
            if (
                record.session_id != session_id
                or record.view_segment_id != view_segment_id
                or record.visual_history_generation != visual_history_generation
                or record.basis_bundle_id != basis_bundle_id
            ):
                continue
            optional_matches = (
                (capture_backend, record.capture_backend),
                (source_hwnd, record.source_hwnd),
                (capture_geometry_id, record.capture_geometry_id),
                (camera_width, record.camera_width),
                (camera_height, record.camera_height),
            )
            if all(expected is None or expected == actual
                   for expected, actual in optional_matches):
                matches.append(record)
        if not matches:
            return None
        return max(matches, key=lambda record: record.created_at_utc)

    def record_live_label(
        self,
        elapsed_s: float,
        weights: Mapping[str, float],
        calibration: Optional[CalibrationRecord] = None,
        snapshot: Optional[RheedFrameSnapshot] = None,
        pyro_temp: Optional[float] = None,
        voltage_V: Optional[float] = None,
        current_A: Optional[float] = None,
        psu_source: str = "none",
        frame: Optional[np.ndarray] = None,
        capture_metadata: Optional[dict] = None,
        equalizer_payload: Optional[Mapping[str, object]] = None,
        labeler: str = "",
    ) -> int:
        """Persist a Live Equalizer label and its exact frozen camera frame.

        Ships workstream #4 from the Jul 10 2026 group meeting. The
        Live Equalizer window subscribes to the camera stream and lets
        the grower drive slider weights against the evolving frame;
        Save triggers this method with the current mixture + a snapshot
        of whatever the window was displaying at click time.

        ``weights`` keys follow the 5 equalizer-class labels: ``"1x1"``,
        ``"Tw(2x1)"``, ``"c(6x2)"``, ``"RT13"``, ``"HTR"``. Missing keys
        default to 0.0 (grower may have moved only a subset of sliders).
        Row schema is intentionally distinct from commit_log.csv so
        downstream analysis can tell "monitor-slider label" from "live-
        equalizer label" without a source column — they live in
        different files.

        An accepted, non-invalid calibration and compatible typed snapshot
        are mandatory. Validation finishes before the frame or counter is
        mutated; HTR is blank while that basis class is inactive.
        """
        if not self._live_label_writer or self._session_dir is None:
            return 0
        if not self._recover_live_label_transaction(self._session_dir):
            log.error("Refusing Live-label write while an older WAL is unresolved")
            return 0
        if calibration is None:
            raise TypeError("record_live_label requires an accepted calibration")
        if snapshot is None:
            raise TypeError("record_live_label requires a RheedFrameSnapshot")
        self._validate_equalizer_context(calibration, snapshot)
        journaled = self.get_calibration(calibration.calibration_id)
        if (
            journaled is None
            or journaled.to_json_dict() != calibration.to_json_dict()
        ):
            raise ValueError("Equalizer calibration is not active in the journal")
        capture_columns = self._snapshot_capture_columns(snapshot)
        if not isinstance(weights, Mapping):
            raise TypeError("weights must be a mapping")

        equalizer_columns = self._equalizer_record_columns(
            equalizer_payload,
            labeler=labeler,
            rgb=snapshot.rgb,
            calibration_valid=True,
        )
        if equalizer_payload is not None:
            supplied_final = equalizer_payload.get("final_weights")
            if not isinstance(supplied_final, Mapping):
                raise ValueError("Equalizer payload final_weights are missing")
            for label in calibration.active_classes:
                if abs(
                    float(weights.get(label, 0.0) or 0.0)
                    - float(supplied_final.get(label, 0.0) or 0.0)
                ) > 1e-9:
                    raise ValueError("Equalizer payload and saved weights differ")

        active_classes = tuple(calibration.active_classes)
        active_text = self._active_classes_text(active_classes)
        active_set = set(active_classes)

        def _validated_weight(label: str) -> str:
            if label not in active_set:
                return ""
            try:
                value = float(weights.get(label, 0.0))
            except (TypeError, ValueError) as exc:
                raise ValueError(f"Invalid Equalizer weight for {label}") from exc
            if not math.isfinite(value) or not 0.0 <= value <= 1.0:
                raise ValueError(f"Equalizer weight for {label} must be in [0, 1]")
            return f"{value:.4f}"

        # Flush previous state before deriving the next durable label index.
        # This also makes a WAL recovery decision independent of Python's file
        # buffering if the process is terminated during the new transaction.
        csv_path = self._session_dir / "live_labels.csv"
        try:
            self._live_label_file.flush()
            os.fsync(self._live_label_file.fileno())
            with open(csv_path, newline="", encoding="utf-8") as stream:
                reader = csv.DictReader(stream)
                if tuple(reader.fieldnames or ()) != tuple(self.LIVE_LABEL_FIELDS):
                    raise ValueError("Live-label CSV schema is incompatible")
                existing_rows = list(reader)
            persisted_indices = [
                int(row["label_idx"]) for row in existing_rows
                if str(row.get("label_idx", "")).strip()
            ]
        except (OSError, TypeError, ValueError) as exc:
            log.error("Could not prepare Live-label transaction: %s", exc)
            return 0

        # Validate every scalar before writing the WAL or image.  In
        # particular, legacy frame/metadata cannot override the frozen atomic
        # snapshot.
        idx = max([self._live_label_counter, *persisted_indices]) + 1
        row = {
            "timestamp": datetime.now().isoformat(),
            "elapsed_s": f"{float(elapsed_s):.2f}",
            "label_idx": idx,
            **equalizer_columns,
            "recon_1x1": _validated_weight("1x1"),
            "recon_tw": _validated_weight("Tw(2x1)"),
            "recon_c6x2": _validated_weight("c(6x2)"),
            "recon_rt13": _validated_weight("RT13"),
            "recon_HTR": _validated_weight("HTR"),
            "pyrometer_temp_C": (
                f"{float(pyro_temp):.1f}" if pyro_temp is not None else ""
            ),
            "voltage_V": (
                f"{float(voltage_V):.3f}" if voltage_V is not None else ""
            ),
            "current_A": (
                f"{float(current_A):.3f}" if current_A is not None else ""
            ),
            "psu_source": str(psu_source),
            "frame_path": "",
            **capture_columns,
            "captured_monotonic_ns": snapshot.received_monotonic_ns,
            "session_id": snapshot.session_id,
            "gun_aligned": str(snapshot.gun_aligned),
            "realignment_active": str(snapshot.realignment_active),
            "calibration_id": calibration.calibration_id,
            "basis_bundle_id": calibration.basis_bundle_id,
            "equalizer_active_classes": active_text,
            "view_segment_id": snapshot.view_segment_id,
            "visual_history_generation": snapshot.visual_history_generation,
        }
        _ = frame, capture_metadata

        ts = datetime.now().strftime("%H%M%S_%f")
        final_relative = Path("frames") / f"live_label_{idx:03d}_{ts}.bmp"
        pending_relative = (
            Path("frames") / f".live_label_{idx:03d}_{ts}.pending.bmp"
        )
        path = self._session_dir / final_relative
        pending_path = self._session_dir / pending_relative
        transaction_path = self._live_label_transaction_path(self._session_dir)
        if path.exists() or pending_path.exists() or transaction_path.exists():
            log.error("Live-label transaction paths already exist for index %d", idx)
            return 0
        row["frame_path"] = str(path)

        transaction = {
            "schema_version": self.LIVE_LABEL_TRANSACTION_SCHEMA_VERSION,
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "pending_frame_path": pending_relative.as_posix(),
            "final_frame_path": final_relative.as_posix(),
            "row": row,
        }
        try:
            wal_bytes = json.dumps(
                transaction,
                sort_keys=True,
                separators=(",", ":"),
                allow_nan=False,
            ).encode("utf-8")
        except (TypeError, ValueError) as exc:
            log.error("Could not serialize Live-label transaction: %s", exc)
            return 0
        if not self._atomic_write_bytes(transaction_path, wal_bytes):
            return 0
        if not self._save_rgb_bmp(snapshot.rgb, pending_path):
            self._remove_live_label_transaction_files(
                pending_path, path, transaction_path,
            )
            return 0
        try:
            # The final name may be temporarily visible before its CSV row,
            # but the durable WAL makes that state explicit and recoverable.
            os.replace(pending_path, path)
        except OSError as exc:
            log.error("Could not finalize Live-label image %s: %s", path, exc)
            self._remove_live_label_transaction_files(
                pending_path, path, transaction_path,
            )
            return 0

        self._close_live_label_stream()
        write_ok = False
        try:
            write_ok = self._atomic_write_csv(
                csv_path,
                self.LIVE_LABEL_FIELDS,
                [*existing_rows, row],
            )
        finally:
            # A BaseException used to model sudden process death deliberately
            # leaves the WAL and image in place; reopening only keeps the test
            # object and recoverable in-process failures usable.
            reopened = self._open_live_label_stream_for_append()
        if not write_ok:
            self._remove_live_label_transaction_files(
                pending_path, path, transaction_path,
            )
            return 0

        self._live_label_counter = idx
        try:
            transaction_path.unlink()
        except OSError as exc:
            # CSV + final image are already committed.  Retaining the WAL is
            # safe: startup recovery verifies the exact row and then clears it.
            log.error(
                "Committed Live label %d but could not clear WAL %s: %s",
                idx,
                transaction_path,
                exc,
            )
        if not reopened:
            log.error("Live-label CSV committed but its append stream is unavailable")
        return idx

    def log_heartbeat(
        self,
        elapsed_s: float,
        pyro_temp: Optional[float] = None,
        frame_path: str = "",
        capture_metadata: Optional[dict] = None,
    ):
        """Append a row to heartbeat_log.csv. Pairs with save_heartbeat_frame."""
        if not self._heartbeat_writer:
            return
        self._heartbeat_writer.writerow({
            "timestamp": datetime.now().isoformat(),
            "elapsed_s": f"{elapsed_s:.2f}",
            "heartbeat_idx": self._heartbeat_counter,
            "pyrometer_temp_C": (
                f"{pyro_temp:.1f}" if pyro_temp is not None else ""
            ),
            "frame_path": frame_path,
            **self._capture_columns(capture_metadata),
        })
        self._heartbeat_file.flush()

    @staticmethod
    def _stringify_csv_row(
        row: Mapping[str, object], fields: Iterable[str],
    ) -> dict[str, str]:
        return {
            field: "" if row.get(field) is None else str(row.get(field))
            for field in fields
        }

    def _build_auto_capture_row(
        self,
        event_idx: int,
        score: float,
        elapsed_s: float,
        pyro_temp: Optional[float] = None,
        buffer_count: int = 0,
        buffer_dir: str = "",
        event_state: str = EVENT_STATE_PENDING,
        capture_metadata: Optional[dict] = None,
        timestamp: Optional[str] = None,
    ) -> dict[str, str]:
        """Validate and stringify one row exactly as CSV reads it back."""
        event_idx = int(event_idx)
        score = float(score)
        elapsed_s = float(elapsed_s)
        buffer_count = int(buffer_count)
        if event_idx <= 0:
            raise ValueError("auto-capture event_idx must be positive")
        if not math.isfinite(score) or not math.isfinite(elapsed_s):
            raise ValueError("auto-capture score and elapsed time must be finite")
        if pyro_temp is not None and not math.isfinite(float(pyro_temp)):
            raise ValueError("auto-capture pyrometer value must be finite")
        if buffer_count < 0:
            raise ValueError("auto-capture buffer_count cannot be negative")
        if event_state not in {
            EVENT_STATE_PENDING,
            EVENT_STATE_KEPT_EXPLICIT,
            EVENT_STATE_KEPT_DEFAULT,
            EVENT_STATE_DISCARDED,
            EVENT_STATE_AUTO_SKIPPED,
        }:
            raise ValueError("invalid auto-capture event state")
        stamp = timestamp or datetime.now().isoformat()
        raw = {
            "timestamp": stamp,
            "elapsed_s": f"{elapsed_s:.2f}",
            "event_idx": event_idx,
            "change_score": f"{score:.4f}",
            "pyrometer_temp_C": (
                f"{float(pyro_temp):.1f}" if pyro_temp is not None else ""
            ),
            "buffer_count": buffer_count,
            "buffer_dir": str(buffer_dir),
            "event_state": event_state,
            "state_changed_at": (
                stamp if event_state != EVENT_STATE_PENDING else ""
            ),
            **self._capture_columns(capture_metadata),
        }
        return self._stringify_csv_row(raw, self.AUTO_CAPTURE_FIELDS)

    def _read_auto_capture_rows(self) -> tuple[Path, list[dict[str, str]]]:
        """Read durable CSV rows, flushing an append stream when available."""
        if self._session_dir is None:
            raise OSError("auto-capture session is not available")
        csv_path = self._session_dir / "auto_capture_events.csv"
        if self._auto_capture_file is not None:
            self._auto_capture_file.flush()
            os.fsync(self._auto_capture_file.fileno())
        with open(csv_path, newline="", encoding="utf-8") as stream:
            reader = csv.DictReader(stream)
            if tuple(reader.fieldnames or ()) != tuple(self.AUTO_CAPTURE_FIELDS):
                raise ValueError("auto-capture CSV schema is incompatible")
            rows = [
                self._stringify_csv_row(row, self.AUTO_CAPTURE_FIELDS)
                for row in reader
            ]
        return csv_path, rows

    def _rollback_auto_capture_transaction(
        self,
        staging_dir: Optional[Path],
        final_dir: Optional[Path],
        marker: Path,
    ) -> bool:
        cleaned = True
        for owned_dir in (staging_dir, final_dir):
            if owned_dir is not None:
                cleaned = self._remove_auto_capture_directory(owned_dir) and cleaned
        if not cleaned:
            log.error("Retaining auto-capture WAL after incomplete rollback: %s", marker)
            return False
        try:
            marker.unlink(missing_ok=True)
            return True
        except OSError as exc:
            log.error("Could not clear rolled-back auto-capture WAL %s: %s", marker, exc)
            return False

    def record_auto_capture_event(
        self,
        *,
        event_idx: int,
        score: float,
        elapsed_s: float,
        frames: list[np.ndarray],
        frame_capture_metadata: Optional[list[dict]] = None,
        pyro_temp: Optional[float] = None,
        event_capture_metadata: Optional[dict] = None,
        classifier_snapshot: Optional[Mapping[str, object]] = None,
    ) -> AutoCaptureCommitResult:
        """Crash-transactionally persist one event row and its frame buffer."""
        if self._session_dir is None or not self._auto_capture_stream_ready():
            return AutoCaptureCommitResult(
                writer_ready=self._auto_capture_stream_ready(),
            )
        try:
            self._auto_capture_file.flush()
            os.fsync(self._auto_capture_file.fileno())
        except (OSError, ValueError) as exc:
            log.error("Could not flush auto-capture CSV before recovery: %s", exc)
            return AutoCaptureCommitResult(
                writer_ready=self._auto_capture_stream_ready(),
            )
        if not self._recover_auto_capture_transaction(self._session_dir):
            log.error("Refusing auto-capture write while an older WAL is unresolved")
            return AutoCaptureCommitResult(
                writer_ready=self._auto_capture_stream_ready(),
            )

        event_idx = int(event_idx)
        if event_idx <= 0:
            raise ValueError("auto-capture event_idx must be positive")
        try:
            from drivers.frame_quality import check_frame_quality
        except ImportError:
            check_frame_quality = None

        ts_tag = datetime.now().strftime("%H%M%S_%f")
        selected: list[tuple[np.ndarray, dict[str, str]]] = []
        for pos, frame in enumerate(frames):
            if check_frame_quality is not None:
                qa = check_frame_quality(frame)
                if not qa.passed:
                    continue
            metadata = (
                frame_capture_metadata[pos]
                if frame_capture_metadata is not None
                and pos < len(frame_capture_metadata)
                else {}
            )
            raw_manifest = {
                "frame_path": f"buf_{pos:02d}_{ts_tag}.bmp",
                **self._capture_columns(metadata),
                "captured_monotonic_ns": metadata.get("captured_monotonic_ns", ""),
                "camera_width": metadata.get("camera_width", frame.shape[1]),
                "camera_height": metadata.get("camera_height", frame.shape[0]),
                "session_id": metadata.get("session_id", ""),
                "view_segment_id": metadata.get("view_segment_id", ""),
                "visual_history_generation": metadata.get(
                    "visual_history_generation", "",
                ),
                "gun_aligned": metadata.get("gun_aligned", ""),
                "realignment_active": metadata.get("realignment_active", ""),
                "calibration_id": metadata.get("calibration_id", ""),
                "basis_bundle_id": metadata.get("basis_bundle_id", ""),
            }
            selected.append((
                frame,
                self._stringify_csv_row(
                    raw_manifest, self.AUTO_CAPTURE_MANIFEST_FIELDS,
                ),
            ))

        buffer_count = len(selected)
        classifier_snapshot_bytes: Optional[bytes] = None
        classifier_snapshot_text = ""
        classifier_snapshot_hash = ""
        if buffer_count and classifier_snapshot is not None:
            if not isinstance(classifier_snapshot, Mapping):
                raise ValueError("classifier snapshot must be a mapping")
            snapshot_payload = dict(classifier_snapshot)
            if int(snapshot_payload.get("event_idx", 0)) != event_idx:
                raise ValueError("classifier snapshot event identity does not match")
            snapshot_score = float(snapshot_payload.get("event_score", math.nan))
            if not math.isfinite(snapshot_score) or snapshot_score != float(score):
                raise ValueError("classifier snapshot score identity does not match")
            try:
                classifier_snapshot_text = json.dumps(
                    snapshot_payload,
                    sort_keys=True,
                    separators=(",", ":"),
                    allow_nan=False,
                )
            except (TypeError, ValueError) as exc:
                raise ValueError("classifier snapshot is not finite JSON") from exc
            classifier_snapshot_bytes = classifier_snapshot_text.encode("utf-8")
            classifier_snapshot_hash = hashlib.sha256(
                classifier_snapshot_bytes,
            ).hexdigest()
        final_relative = (
            Path("frames") / f"auto_event_{event_idx:03d}"
            if buffer_count else None
        )
        staging_relative = (
            Path("frames")
            / f".auto_event_{event_idx:03d}.{uuid.uuid4().hex}.pending"
            if buffer_count else None
        )
        buffer_dir = final_relative.as_posix() if final_relative is not None else ""
        row = self._build_auto_capture_row(
            event_idx=event_idx,
            score=score,
            elapsed_s=elapsed_s,
            pyro_temp=pyro_temp,
            buffer_count=buffer_count,
            buffer_dir=buffer_dir,
            event_state=(
                EVENT_STATE_PENDING if buffer_count else EVENT_STATE_AUTO_SKIPPED
            ),
            capture_metadata=event_capture_metadata,
        )
        manifest_rows = [manifest for _frame, manifest in selected]
        try:
            csv_path, existing_rows = self._read_auto_capture_rows()
        except (OSError, TypeError, ValueError) as exc:
            log.error("Could not prepare auto-capture transaction: %s", exc)
            return AutoCaptureCommitResult(
                writer_ready=self._auto_capture_stream_ready(),
            )
        if any(row_["event_idx"] == str(event_idx) for row_ in existing_rows):
            log.error("Auto-capture event index %d already exists", event_idx)
            return AutoCaptureCommitResult(
                writer_ready=self._auto_capture_stream_ready(),
            )

        marker = self._auto_capture_transaction_path(self._session_dir)
        staging_dir = (
            self._session_dir / staging_relative
            if staging_relative is not None else None
        )
        final_dir = (
            self._session_dir / final_relative
            if final_relative is not None else None
        )
        expected_final_dir = (
            self._session_dir / "frames" / f"auto_event_{event_idx:03d}"
        )
        if marker.exists() or any(
            path is not None and path.exists() for path in (staging_dir, final_dir)
        ) or expected_final_dir.exists():
            log.error("Auto-capture transaction paths already exist for event %d", event_idx)
            return AutoCaptureCommitResult(
                writer_ready=self._auto_capture_stream_ready(),
            )
        wal_payload = {
            "schema_version": self.AUTO_CAPTURE_TRANSACTION_SCHEMA_VERSION,
            "event_idx": event_idx,
            "created_at_utc": datetime.now(timezone.utc).isoformat(),
            "staging_dir": (
                staging_relative.as_posix() if staging_relative is not None else ""
            ),
            "final_dir": final_relative.as_posix() if final_relative is not None else "",
            "row": row,
            "manifest_rows": manifest_rows,
            "classifier_snapshot_path": (
                "classifier_state.json"
                if classifier_snapshot_bytes is not None else ""
            ),
            "classifier_snapshot_json": classifier_snapshot_text,
            "classifier_snapshot_sha256": classifier_snapshot_hash,
        }
        try:
            wal_bytes = json.dumps(
                wal_payload,
                sort_keys=True,
                separators=(",", ":"),
                allow_nan=False,
            ).encode("utf-8")
        except (TypeError, ValueError) as exc:
            log.error("Could not serialize auto-capture transaction: %s", exc)
            return AutoCaptureCommitResult(
                writer_ready=self._auto_capture_stream_ready(),
            )
        if not self._atomic_write_bytes(marker, wal_bytes):
            return AutoCaptureCommitResult(
                writer_ready=self._auto_capture_stream_ready(),
            )

        if buffer_count:
            assert staging_dir is not None and final_dir is not None
            try:
                staging_dir.mkdir()
            except OSError as exc:
                log.error("Could not create auto-capture staging directory: %s", exc)
                self._rollback_auto_capture_transaction(staging_dir, final_dir, marker)
                return AutoCaptureCommitResult(
                    writer_ready=self._auto_capture_stream_ready(),
                )
            for frame, manifest in selected:
                if not self._save_rgb_bmp(frame, staging_dir / manifest["frame_path"]):
                    self._rollback_auto_capture_transaction(staging_dir, final_dir, marker)
                    return AutoCaptureCommitResult(
                        writer_ready=self._auto_capture_stream_ready(),
                    )
            if (
                classifier_snapshot_bytes is not None
                and not self._atomic_write_bytes(
                    staging_dir / "classifier_state.json",
                    classifier_snapshot_bytes,
                )
            ):
                self._rollback_auto_capture_transaction(staging_dir, final_dir, marker)
                return AutoCaptureCommitResult(
                    writer_ready=self._auto_capture_stream_ready(),
                )
            if not self._atomic_write_csv(
                staging_dir / "capture_manifest.csv",
                self.AUTO_CAPTURE_MANIFEST_FIELDS,
                manifest_rows,
            ):
                self._rollback_auto_capture_transaction(staging_dir, final_dir, marker)
                return AutoCaptureCommitResult(
                    writer_ready=self._auto_capture_stream_ready(),
                )
            try:
                os.replace(staging_dir, final_dir)
            except OSError as exc:
                log.error("Could not finalize auto-capture event directory: %s", exc)
                self._rollback_auto_capture_transaction(staging_dir, final_dir, marker)
                return AutoCaptureCommitResult(
                    writer_ready=self._auto_capture_stream_ready(),
                )

        self._close_auto_capture_stream()
        write_ok = False
        try:
            write_ok = self._atomic_write_csv(
                csv_path,
                self.AUTO_CAPTURE_FIELDS,
                [*existing_rows, row],
            )
        finally:
            reopened = self._open_auto_capture_stream_for_append()
        if not write_ok:
            self._rollback_auto_capture_transaction(staging_dir, final_dir, marker)
            return AutoCaptureCommitResult(writer_ready=bool(reopened))
        try:
            marker.unlink()
        except OSError as exc:
            log.error(
                "Committed auto-capture event %d but could not clear WAL %s: %s",
                event_idx,
                marker,
                exc,
            )
        if not reopened:
            log.error("Auto-capture CSV committed but its append stream is unavailable")
        return AutoCaptureCommitResult(
            buffer_count=buffer_count,
            buffer_dir=buffer_dir,
            committed=True,
            writer_ready=bool(reopened and self._auto_capture_stream_ready()),
        )

    def log_auto_capture_event(
        self,
        event_idx: int,
        score: float,
        elapsed_s: float,
        pyro_temp: Optional[float] = None,
        buffer_count: int = 0,
        buffer_dir: str = "",
        event_state: str = EVENT_STATE_PENDING,
        capture_metadata: Optional[dict] = None,
    ) -> bool:
        """Atomically append a no-image row to auto_capture_events.csv.

        Called by GrowthApp when AutoCaptureEngine emits frame_captured.
        Writes timestamp + change score + temp at trigger time, so the
        flagged moments can be cross-referenced against grower notes
        and pyrometer trajectory after the session. ``buffer_count`` and
        ``buffer_dir`` capture how many context frames were dumped and
        where, so post-hoc analysis can locate them.

        ``event_state`` defaults to ``"pending"`` at fire time and is updated
        to one of ``"kept_explicit"`` / ``"kept_default"`` / ``"discarded"``
        when the grower interacts with the AutoCaptureBanner (see
        ``update_auto_capture_state``).
        """
        if not self._auto_capture_writer or self._session_dir is None:
            return False
        if int(buffer_count) != 0 or str(buffer_dir):
            raise ValueError(
                "Buffered auto-capture events require record_auto_capture_event"
            )
        if not self._recover_auto_capture_transaction(self._session_dir):
            return False
        row = self._build_auto_capture_row(
            event_idx=event_idx,
            score=score,
            elapsed_s=elapsed_s,
            pyro_temp=pyro_temp,
            buffer_count=0,
            buffer_dir="",
            event_state=event_state,
            capture_metadata=capture_metadata,
        )
        try:
            csv_path, rows = self._read_auto_capture_rows()
        except (OSError, TypeError, ValueError) as exc:
            log.error("Could not prepare auto-capture row: %s", exc)
            return False
        if any(existing["event_idx"] == row["event_idx"] for existing in rows):
            return False
        self._close_auto_capture_stream()
        try:
            write_ok = self._atomic_write_csv(
                csv_path, self.AUTO_CAPTURE_FIELDS, [*rows, row],
            )
        finally:
            self._open_auto_capture_stream_for_append()
        return write_ok

    def update_auto_capture_state(
        self,
        event_idx: int,
        new_state: str,
    ) -> bool:
        """Update event_state and state_changed_at for an existing event row.

        Called when the grower interacts with the AutoCaptureBanner (Keep
        Now, Discard) or when the keep-default countdown fires. Rewrites
        auto_capture_events.csv in place — small file, infrequent updates,
        simple semantics. Returns True if the row was found and updated.

        Non-destructive design: discard updates the row state but does NOT
        remove the buffer directory. The grower can recover discarded
        events from the Events tab if they change their mind. See
        ``feedback_aiqm_grower_friction.md`` for the design rationale.
        """
        if self._session_dir is None:
            return False
        try:
            if self._auto_capture_file is not None:
                self._auto_capture_file.flush()
                os.fsync(self._auto_capture_file.fileno())
        except (OSError, ValueError):
            return False
        if not self._recover_auto_capture_transaction(self._session_dir):
            return False
        try:
            csv_path, rows = self._read_auto_capture_rows()
        except (OSError, TypeError, ValueError):
            return False

        # Close the writer's file handle so we can rewrite in-place.
        self._close_auto_capture_stream()

        timestamp = datetime.now().isoformat()
        found = False
        for row in rows:
            if str(row.get("event_idx", "")) == str(event_idx):
                row["event_state"] = new_state
                row["state_changed_at"] = timestamp
                found = True
                break

        write_ok = True
        if found:
            write_ok = self._atomic_write_csv(
                csv_path,
                self.AUTO_CAPTURE_FIELDS,
                rows,
            )

        # Reopen for append so subsequent log_auto_capture_event calls work.
        # No header write — the file already has the header from start_session
        # (or the rewrite above when found=True).
        self._open_auto_capture_stream_for_append()
        return found and write_ok

    def read_event_labels(self) -> dict[int, dict]:
        """Load all rows from events_labels.csv keyed by event_idx.

        Returns an empty dict if no session is active, the file doesn't
        exist yet (no labels applied), or the file can't be read. The
        Events tab uses this once per session attach to seed an in-memory
        cache; the cache is then kept in sync as the grower applies new
        labels through update_event_label.
        """
        if self._session_dir is None:
            return {}
        csv_path = self._session_dir / "events_labels.csv"
        if not csv_path.exists():
            return {}
        labels: dict[int, dict] = {}
        try:
            with open(csv_path, "r", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        idx = int(row.get("event_idx", "") or 0)
                    except (TypeError, ValueError):
                        continue
                    labels[idx] = {
                        field: row.get(field, "")
                        for field in self.EVENT_LABEL_FIELDS
                    }
        except OSError:
            return {}
        return labels

    def update_event_label(
        self,
        event_idx: int,
        primary_reconstruction: Optional[str] = None,
        human_primary_reconstruction: Optional[str] = None,
        human_label_source: Optional[str] = None,
        human_labeler: Optional[str] = None,
        human_confidence: object = None,
        human_blind_to_model: Optional[bool] = None,
        human_blind_to_equalizer: Optional[bool] = None,
        human_frame_path: Optional[str] = None,
        human_capture_metadata: Optional[Mapping[str, object]] = None,
        human_annotation_id: str = "",
        change_from: Optional[str] = None,
        change_to: Optional[str] = None,
        notes: Optional[str] = None,
        recon_1x1: Optional[float] = None,
        recon_tw: Optional[float] = None,
        recon_c6x2: Optional[float] = None,
        recon_rt13: Optional[float] = None,
        recon_HTR: Optional[float] = None,
        calibration: Optional[CalibrationRecord] = None,
        snapshot: Optional[RheedFrameSnapshot] = None,
        calibration_id: Optional[str] = None,
        basis_bundle_id: Optional[str] = None,
        equalizer_active_classes: Optional[Iterable[str]] = None,
        view_segment_id: Optional[int] = None,
        visual_history_generation: Optional[int] = None,
        frame_path: Optional[str] = None,
        capture_backend: Optional[str] = None,
        captured_at_utc: Optional[str] = None,
        capture_sequence: Optional[int] = None,
        frame_age_ms: Optional[float] = None,
        source_hwnd: Optional[int] = None,
        capture_geometry_id: Optional[str] = None,
        equalizer_payload: Optional[Mapping[str, object]] = None,
        equalizer_labeler: str = "",
    ) -> bool:
        """Atomically upsert a labeling row in events_labels.csv.

        Reads the existing file (if any), updates or appends the row for
        ``event_idx``, and rewrites the whole file. Each call is a
        per-change atomic write — the design accepts ~3-5 rewrites per
        labeled event (one per dropdown / notes commit) in exchange for
        no "save button" cognitive load on the grower.

        Only fields explicitly passed are updated; ``None`` means "leave
        existing value alone." This lets the EventsTab call with just
        ``primary_reconstruction=...`` when the dropdown changes without
        clobbering a previously-typed notes string. ``label_timestamp_iso``
        is always refreshed to record when the label was last touched.

        Kwarg source map (Jul 15 2026):
          - primary_reconstruction, change_from, change_to → three
            dropdowns on events_tab labeling form (Jul 15). change_from/to
            went from "reserved-but-unused" to UI-populated in the same
            commit that landed the labeling-form dropdowns.
          - notes → events_tab text field (debounced)
          - recon_1x1 / recon_tw / recon_c6x2 / recon_rt13 / recon_HTR →
            Equalizer popup save callback (see events_tab.
            _make_equalizer_save_callback)

        File is created on first write — sessions with no labeling
        activity won't leave behind an empty events_labels.csv.

        Returns True on successful write, False on no-session or I/O error.
        """
        if self._session_dir is None:
            return False
        csv_path = self._session_dir / "events_labels.csv"

        recon_values = {
            "recon_1x1": recon_1x1,
            "recon_tw": recon_tw,
            "recon_c6x2": recon_c6x2,
            "recon_rt13": recon_rt13,
            "recon_HTR": recon_HTR,
        }
        has_equalizer_weights = any(
            value is not None for value in recon_values.values()
        )
        if equalizer_payload is not None and not has_equalizer_weights:
            raise ValueError("Equalizer payload requires final weights")
        has_typed_context = calibration is not None or snapshot is not None
        has_explicit_context = any(value is not None for value in (
            calibration_id, basis_bundle_id, equalizer_active_classes,
            view_segment_id, visual_history_generation, frame_path,
            capture_backend, captured_at_utc, capture_sequence,
            frame_age_ms, source_hwnd, capture_geometry_id,
        ))

        context: dict[str, object] = {}
        if has_typed_context:
            if calibration is None or snapshot is None:
                raise TypeError("calibration and snapshot must be supplied together")
            self._validate_equalizer_context(calibration, snapshot)
            journaled = self.get_historical_calibration(
                calibration.calibration_id,
            )
            if (
                journaled is None
                or journaled.to_json_dict() != calibration.to_json_dict()
            ):
                raise ValueError(
                    "Equalizer calibration is not present in the trusted journal"
                )
            capture = self._snapshot_capture_columns(snapshot)
            active_text = self._active_classes_text(calibration.active_classes)
            existing_frame_path = self._resolve_existing_event_frame(
                self._session_dir,
                frame_path or "",
            )
            context = {
                "calibration_id": calibration.calibration_id,
                "basis_bundle_id": calibration.basis_bundle_id,
                "equalizer_active_classes": active_text,
                "view_segment_id": snapshot.view_segment_id,
                "visual_history_generation": snapshot.visual_history_generation,
                "frame_path": str(existing_frame_path),
                "captured_monotonic_ns": snapshot.received_monotonic_ns,
                "gun_aligned": str(snapshot.gun_aligned),
                "realignment_active": str(snapshot.realignment_active),
                "session_id": snapshot.session_id,
                **capture,
            }
            active_classes = set(calibration.active_classes)
        elif has_equalizer_weights or has_explicit_context:
            # The Events view may already have a frame on disk.  In that case
            # accept explicit CSV provenance only when its calibration ID
            # resolves to an accepted historical journal record.
            if not calibration_id:
                raise ValueError("Explicit Equalizer context requires calibration_id")
            journaled = self.get_historical_calibration(calibration_id)
            if journaled is None:
                raise ValueError("Calibration ID is missing from the trusted journal")
            resolved_bundle = basis_bundle_id or journaled.basis_bundle_id
            resolved_active = (
                equalizer_active_classes
                if equalizer_active_classes is not None
                else journaled.active_classes
            )
            active_text = self._active_classes_text(resolved_active)
            active_classes = set(json.loads(active_text))
            required = {
                "view_segment_id": view_segment_id,
                "visual_history_generation": visual_history_generation,
                "frame_path": frame_path,
                "capture_backend": capture_backend,
                "captured_at_utc": captured_at_utc,
                "capture_sequence": capture_sequence,
                "frame_age_ms": frame_age_ms,
                "source_hwnd": source_hwnd,
                "capture_geometry_id": capture_geometry_id,
            }
            missing = [name for name, value in required.items() if value is None]
            if missing:
                raise ValueError(
                    "Explicit Equalizer context is missing: " + ", ".join(missing)
                )
            existing_frame_path = self._resolve_existing_event_frame(
                self._session_dir,
                str(frame_path),
            )
            if (
                resolved_bundle != journaled.basis_bundle_id
                or view_segment_id != journaled.view_segment_id
                or visual_history_generation != journaled.visual_history_generation
                or capture_backend != journaled.capture_backend
                or source_hwnd != journaled.source_hwnd
                or capture_geometry_id != journaled.capture_geometry_id
            ):
                raise ValueError("Explicit Equalizer context is incompatible")
            age = float(frame_age_ms)
            if not math.isfinite(age) or age < 0:
                raise ValueError("Explicit frame_age_ms must be finite and non-negative")
            context = {
                "calibration_id": journaled.calibration_id,
                "basis_bundle_id": journaled.basis_bundle_id,
                "equalizer_active_classes": active_text,
                "view_segment_id": view_segment_id,
                "visual_history_generation": visual_history_generation,
                "frame_path": str(existing_frame_path),
                "captured_monotonic_ns": journaled.received_monotonic_ns,
                "gun_aligned": str(journaled.gun_aligned),
                "realignment_active": str(journaled.realignment_active),
                "session_id": journaled.session_id,
                **self._capture_columns({
                    "capture_backend": capture_backend,
                    "captured_at_utc": captured_at_utc,
                    "capture_sequence": capture_sequence,
                    "frame_age_ms": age,
                    "source_hwnd": source_hwnd,
                    "capture_geometry_id": capture_geometry_id,
                }),
            }
        else:
            active_classes = set()

        equalizer_columns: dict[str, object] = {}
        if has_equalizer_weights:
            equalizer_columns = self._equalizer_record_columns(
                equalizer_payload,
                labeler=equalizer_labeler,
                rgb=snapshot.rgb if snapshot is not None else None,
                calibration_valid=True,
            )
            if equalizer_payload is not None:
                supplied_final = equalizer_payload.get("final_weights")
                if not isinstance(supplied_final, Mapping):
                    raise ValueError("Equalizer payload final_weights are missing")
                payload_columns = {
                    "recon_1x1": "1x1",
                    "recon_tw": "Tw(2x1)",
                    "recon_c6x2": "c(6x2)",
                    "recon_rt13": "RT13",
                }
                for column, label in payload_columns.items():
                    value = recon_values[column]
                    if value is None or abs(
                        float(value) - float(supplied_final.get(label, 0.0) or 0.0)
                    ) > 1e-9:
                        raise ValueError("Equalizer payload and saved weights differ")

        human_columns: dict[str, object] = {}
        if human_primary_reconstruction is not None:
            if not human_primary_reconstruction:
                human_columns = {
                    field: ""
                    for field in self.EVENT_LABEL_FIELDS
                    if field.startswith("human_")
                }
            else:
                if human_frame_path is None or human_capture_metadata is None:
                    raise ValueError("Human primary label requires exact frame context")
                human_columns = self._human_primary_columns(
                    reconstruction=human_primary_reconstruction,
                    label_source=str(human_label_source or ""),
                    labeler=str(human_labeler or ""),
                    confidence=human_confidence,
                    blind_to_model=bool(human_blind_to_model),
                    blind_to_equalizer=bool(human_blind_to_equalizer),
                    frame_path=Path(human_frame_path),
                    capture_metadata=human_capture_metadata,
                    session_dir=self._session_dir,
                )

        formatted_recons: dict[str, str] = {}
        for column, value in recon_values.items():
            if value is None:
                continue
            label = {
                "recon_1x1": "1x1",
                "recon_tw": "Tw(2x1)",
                "recon_c6x2": "c(6x2)",
                "recon_rt13": "RT13",
                "recon_HTR": "HTR",
            }[column]
            if label not in active_classes:
                formatted_recons[column] = ""
                continue
            numeric = float(value)
            if not math.isfinite(numeric) or not 0.0 <= numeric <= 1.0:
                raise ValueError(f"{column} must be in [0, 1]")
            formatted_recons[column] = f"{numeric:.4f}"

        rows: list[dict] = []
        if csv_path.exists():
            try:
                with open(csv_path, "r", newline="") as f:
                    reader = csv.DictReader(f)
                    rows = [
                        {
                            field: r.get(field, "")
                            for field in self.EVENT_LABEL_FIELDS
                        }
                        for r in reader
                    ]
            except OSError:
                return False

        target = str(event_idx)
        existing = next(
            (r for r in rows if str(r.get("event_idx", "")) == target),
            None,
        )
        if existing is None:
            existing = {f: "" for f in self.EVENT_LABEL_FIELDS}
            existing["event_idx"] = target
            rows.append(existing)

        label_timestamp = self._utc_now_text()
        human_audit_row: Optional[dict[str, str]] = None
        if human_primary_reconstruction:
            # The append-only file is the source of truth.  It is committed
            # before the mutable per-event summary so a crash cannot leave a
            # summary-only "gold" label with no independent annotation.
            human_audit_row = self._append_human_primary_label(
                human_columns=human_columns,
                capture_metadata=human_capture_metadata or {},
                annotation_id=human_annotation_id,
                timestamp=label_timestamp,
            )

        if primary_reconstruction is not None:
            existing["primary_reconstruction"] = primary_reconstruction
        if human_primary_reconstruction is not None:
            # Legacy summary mirrors only an explicit human decision.  An
            # Equalizer save never reaches this branch.
            existing["primary_reconstruction"] = human_primary_reconstruction
            existing.update(human_columns)
            if human_audit_row is not None:
                existing["human_primary_annotation_id"] = human_audit_row[
                    "annotation_id"
                ]
                existing["human_primary_label_idx"] = human_audit_row[
                    "label_idx"
                ]
        if change_from is not None:
            existing["change_from"] = change_from
        if change_to is not None:
            existing["change_to"] = change_to
        if notes is not None:
            existing["notes"] = notes
        existing.update(formatted_recons)
        existing.update(equalizer_columns)
        if context:
            existing.update(context)
            if "HTR" not in active_classes:
                existing["recon_HTR"] = ""
        existing["label_timestamp_iso"] = label_timestamp

        if not self._atomic_write_csv(
            csv_path,
            self.EVENT_LABEL_FIELDS,
            rows,
        ):
            if human_audit_row is not None:
                log.error(
                    "Human annotation %s is durable but events_labels.csv "
                    "summary could not be refreshed",
                    human_audit_row["annotation_id"],
                )
                return True
            return False
        return True

    def save_auto_capture_buffer(
        self,
        event_idx: int,
        frames: list[np.ndarray],
        capture_metadata: Optional[list[dict]] = None,
    ) -> tuple[int, str]:
        """Fail closed for callers of the removed two-step buffer API."""
        log.error(
            "save_auto_capture_buffer is disabled; use the transactional "
            "record_auto_capture_event API"
        )
        return 0, ""

    def save_frame(
        self, frame: np.ndarray, timestamp: str = "",
    ) -> tuple[str, Optional[bool]]:
        """Save frame as PNG to session frames/ subdir. Returns (path, quality_pass).

        Manual LOG ENTRY intent takes precedence over the frame quality
        gate — if the grower explicitly clicked LOG ENTRY, the frame is
        saved regardless of quality. The gate result is captured as
        metadata (quality_pass) so downstream training-data consumers
        can filter to high-quality frames while the grower's original
        intent is preserved in the archive.

        This is a deliberate contract difference from save_heartbeat_frame
        and save_auto_capture_buffer (both periodic + implicit — quality
        gate rejects there because no grower is asking for that frame).

        Returns:
            (path, quality_pass) tuple.
              path: str filesystem path if the frame was saved successfully,
                    "" if no session dir or PIL/cv2 both missing.
              quality_pass: True if the quality gate passed, False if it
                    flagged the frame (still saved), None if the gate
                    couldn't be evaluated (frame_quality module missing).
        """
        if self._session_dir is None:
            return "", None

        # Capture the quality gate result but don't gate saving on it.
        # LOG ENTRY = grower's explicit intent (see Jul 8 2026 decision
        # in bulbasaur_lab_day_jul07.md follow-up 2).
        quality_pass: Optional[bool] = None
        try:
            from drivers.frame_quality import check_frame_quality
            qa = check_frame_quality(frame)
            quality_pass = bool(qa.passed)
            if not qa.passed:
                import sys
                print(
                    f"[GrowthLogger] LOG ENTRY frame quality flagged "
                    f"(saved anyway per grower intent): {qa.reason}",
                    file=sys.stderr,
                    flush=True,
                )
        except ImportError:
            pass  # Quality gate optional; leave quality_pass=None.

        self._commit_counter += 1
        ts = timestamp or datetime.now().strftime("%H%M%S")
        # .bmp per Jul 9 2026 switch — matches Justin's training data
        # format. Files ~10x larger than PNG (~300 KB vs ~30 KB for a
        # 656x492x3 uint8 frame), still trivial vs the T9 SSD's capacity.
        fname = f"entry_{self._commit_counter:03d}_{ts}.bmp"
        path = self._session_dir / "frames" / fname

        try:
            from PIL import Image
            img = Image.fromarray(frame)
            img.save(str(path))
        except ImportError:
            try:
                import cv2
                cv2.imwrite(str(path), cv2.cvtColor(frame, cv2.COLOR_RGB2BGR))
            except ImportError:
                return "", quality_pass

        return str(path), quality_pass

    def save_session_metadata(self, metadata: dict):
        """Save session metadata to a JSON file."""
        if self._session_dir is None:
            return
        now = datetime.now()
        meta = {
            **metadata,
            "session_start": self._session_start.isoformat() if self._session_start else None,
            "session_end": now.isoformat(),
            "session_duration_s": (
                round((now - self._session_start).total_seconds(), 1)
                if self._session_start else None
            ),
            "sensor_row_count": self._sensor_row_counter,
            "commit_entry_count": len(self._entries),
            "heartbeat_frame_count": self._heartbeat_counter,
            "manual_event_count": self._manual_event_counter,
            "rheed_view_event_count": self._rheed_view_event_counter,
            "live_label_count": self._live_label_counter,
            "equalizer_calibration_event_count": (
                self._equalizer_calibration_event_counter
            ),
            "equalizer_basis_bundle_count": len(
                self._basis_bundle_ids_recorded
            ),
        }
        meta_path = self._session_dir / "session_metadata.json"
        with open(meta_path, "w") as f:
            json.dump(meta, f, indent=2)

    def export_growth_log(self, metadata: dict) -> str:
        """Export session as OMBE growth log. Returns file path or empty string."""
        if self._session_dir is None or not self._entries:
            return ""

        try:
            return self._export_xlsx(metadata)
        except ImportError:
            return self._export_csv_log(metadata)

    def _export_xlsx(self, metadata: dict) -> str:
        """Export as xlsx matching OMBE growth log template."""
        from openpyxl import Workbook
        from openpyxl.styles import Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Growth Log"

        bold = Font(bold=True)
        bold_large = Font(bold=True, size=12)

        # --- Header section ---
        ws['B2'] = 'Date:'
        ws['B2'].font = bold
        ws['D2'] = metadata.get('date', '')
        ws['G2'] = 'Substrate:'
        ws['G2'].font = bold

        ws['B3'] = 'Grower:'
        ws['B3'].font = bold
        ws['D3'] = metadata.get('grower', '')
        ws['G3'] = 'Sample ID:'
        ws['G3'].font = bold
        ws['I3'] = metadata.get('sample_id', '')

        ws['B4'] = 'Base pressure (mbar):'
        ws['B4'].font = bold
        ws['G4'] = 'Growth pressure (mbar):'
        ws['G4'].font = bold

        # --- Source parameter headers (OMBE elements) ---
        elements = ['Substrate', 'Sr', 'Ti', 'Y', 'Er', 'Eu', 'O', 'Al', 'Ta']
        for i, elem in enumerate(elements):
            cell = ws.cell(row=5, column=3 + i, value=elem)
            cell.font = bold
        ws.cell(row=5, column=12, value='Flux ratio').font = bold

        ws['B6'] = 'Temperature (\u2103)'
        ws['B6'].font = bold
        ws['B7'] = 'Flux (mbar)'
        ws['B7'].font = bold
        ws['B8'] = 'Time (min)'
        ws['B8'].font = bold

        # --- Growth notes section ---
        ws['B9'] = 'Growth Notes'
        ws['B9'].font = bold_large

        ws['B10'] = 'Time (hh:mm)'
        ws['B10'].font = bold
        ws['C10'] = 'Temp (\u2103)'
        ws['C10'].font = bold
        ws['D10'] = 'Operation'
        ws['D10'].font = bold

        # Pre/post annealing headers
        ws['C11'] = 'Pre-growth annealing T (\u2103)'
        ws['F11'] = 'Pre-growth annealing time (min)'
        ws['H11'] = 'Post-growth annealing T (\u2103)'
        ws['K11'] = 'Post-growth annealing time (min)'

        # --- Operations log entries ---
        for i, entry in enumerate(self._entries):
            row = 13 + i
            ws.cell(row=row, column=2, value=entry.get('time_display', ''))
            temp = entry.get('pyrometer_temp_C', '')
            if temp:
                try:
                    ws.cell(row=row, column=3, value=float(temp))
                except ValueError:
                    ws.cell(row=row, column=3, value=temp)
            ws.cell(row=row, column=4, value=entry.get('note', ''))

        # --- Column widths ---
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['G'].width = 22
        ws.column_dimensions['I'].width = 20

        export_path = self._session_dir / "growth_log.xlsx"
        wb.save(str(export_path))
        return str(export_path)

    def _export_csv_log(self, metadata: dict) -> str:
        """Fallback CSV export if openpyxl is not available."""
        if self._session_dir is None:
            return ""
        export_path = self._session_dir / "growth_log_export.csv"
        with open(export_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["OMBE Growth Log"])
            writer.writerow(["Date", metadata.get('date', '')])
            writer.writerow(["Grower", metadata.get('grower', '')])
            writer.writerow(["Sample ID", metadata.get('sample_id', '')])
            writer.writerow([])
            writer.writerow([
                "Time", "Temp (\u2103)", "Voltage (V)", "Current (A)", "Note",
            ])
            for entry in self._entries:
                writer.writerow([
                    entry.get("time_display", ""),
                    entry.get("pyrometer_temp_C", ""),
                    entry.get("voltage_V", ""),
                    entry.get("current_A", ""),
                    entry.get("note", ""),
                ])
        return str(export_path)

    def end_session(self):
        """Close CSV files. Preserves session_dir and entries for post-stop export."""
        for f in (
            self._sensor_file, self._commit_file,
            self._auto_capture_file, self._heartbeat_file,
            self._set_change_file, self._manual_event_file,
            self._rheed_view_event_file,
            self._live_label_file,
            self._equalizer_calibration_file,
        ):
            if f and not f.closed:
                f.close()
        self._sensor_file = None
        self._sensor_writer = None
        self._commit_file = None
        self._commit_writer = None
        self._auto_capture_file = None
        self._auto_capture_writer = None
        self._heartbeat_file = None
        self._heartbeat_writer = None
        self._set_change_file = None
        self._set_change_writer = None
        self._manual_event_file = None
        self._manual_event_writer = None
        self._rheed_view_event_file = None
        self._rheed_view_event_writer = None
        self._live_label_file = None
        self._live_label_writer = None
        self._equalizer_calibration_file = None
        # NOTE: _session_dir and _entries intentionally preserved
        # so Export Growth Log works after STOP.

    def generate_temperature_plot(self, metadata: Optional[dict] = None) -> Optional[Path]:
        """Generate a Temperature-vs-Time PNG from this session's sensor_log.csv.

        Convenience auto-plot called from GrowthApp at session end (per Jun 23
        2026 — Frankie's f_version pattern, ported to the canonical product).
        Reads ``pyrometer_temp_C`` + ``elapsed_s`` from the sensor log this
        session just wrote and saves ``temperature_profile.png`` to the same
        directory.

        Works post-``end_session()`` because ``_session_dir`` is preserved
        through close. Returns the plot path on success; returns None and
        logs a warning if matplotlib isn't installed or the sensor log has
        no temperature data (e.g. pyrometer disconnected the whole session).

        For richer post-hoc plotting (custom titles, output format, dpi),
        use ``python scripts/plot_temperature.py <session_dir>``.
        """
        if self._session_dir is None:
            return None

        sensor_path = self._session_dir / "sensor_log.csv"
        if not sensor_path.exists():
            return None

        # Matplotlib is a soft dependency — present on Mac dev env, may not
        # be on Bulbasaur. Lazy-import + graceful failure.
        try:
            import matplotlib
            matplotlib.use("Agg")  # No GUI needed; just write a file
            import matplotlib.pyplot as plt
        except ImportError:
            import logging
            logging.getLogger(__name__).warning(
                "matplotlib not installed; skipping auto T-vs-t plot. "
                "Install with: pip install matplotlib"
            )
            return None

        # Parse sensor_log.csv — same shape as scripts/plot_temperature.py
        elapsed: list[float] = []
        temps: list[float] = []
        with open(sensor_path, newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                try:
                    t = float(row["elapsed_s"])
                except (KeyError, ValueError):
                    continue
                temp_str = row.get("pyrometer_temp_C", "").strip()
                if not temp_str:
                    continue
                try:
                    temp = float(temp_str)
                except ValueError:
                    continue
                elapsed.append(t)
                temps.append(temp)

        if not temps:
            # Pyrometer was offline the whole session — nothing to plot.
            return None

        # Title: prefer metadata if provided (richer); fall back to dir name.
        if metadata:
            title_parts = []
            if metadata.get("sample_id"):
                title_parts.append(str(metadata["sample_id"]))
            if metadata.get("grower"):
                title_parts.append(f"by {metadata['grower']}")
            if metadata.get("date"):
                title_parts.append(metadata["date"])
            title = " — ".join(title_parts) if title_parts else self._session_dir.name
        else:
            title = self._session_dir.name

        elapsed_min = [t / 60.0 for t in elapsed]

        fig, ax = plt.subplots(figsize=(10, 5))
        ax.plot(elapsed_min, temps, color="#0d9488", linewidth=1.2)
        ax.set_xlabel("Elapsed Time (min)", fontsize=12)
        ax.set_ylabel("Temperature (°C)", fontsize=12)
        ax.set_title(title, fontsize=13)
        ax.grid(True, alpha=0.3)

        # Annotate min/max so the curve is glanceable
        t_min, t_max = min(temps), max(temps)
        ax.axhline(y=t_max, color="#dc2626", linestyle="--", alpha=0.5, linewidth=0.8)
        ax.axhline(y=t_min, color="#2563eb", linestyle="--", alpha=0.5, linewidth=0.8)
        ax.text(
            elapsed_min[-1], t_max, f"  {t_max:.0f}°C",
            va="bottom", color="#dc2626", fontsize=9,
        )
        ax.text(
            elapsed_min[-1], t_min, f"  {t_min:.0f}°C",
            va="top", color="#2563eb", fontsize=9,
        )

        fig.tight_layout()
        out_path = self._session_dir / "temperature_profile.png"
        fig.savefig(str(out_path), dpi=150)
        plt.close(fig)  # Free figure memory; we're done with it.
        return out_path
